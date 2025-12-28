import sys, os, re, csv, math, hashlib
from dataclasses import dataclass
from typing import List, Dict, Tuple, Optional

import pandas as pd

from PySide6.QtCore import Qt, QThread, Signal
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QFileDialog, QMessageBox,
    QVBoxLayout, QHBoxLayout, QLabel, QPushButton, QListWidget, QListWidgetItem,
    QSpinBox, QDoubleSpinBox, QCheckBox, QLineEdit, QProgressBar, QGroupBox, QFormLayout
)

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


# ---------------------------- Core logic ----------------------------

NOISY_NAME_DEFAULT = r"(status|state|type|category|region|city|country|method|channel|desc|description|note|notes|comment|remarks|text|message|name)$"

def normalize_value(v, case_insensitive=True, trim=True, numeric_canonical=True):
    if v is None or (isinstance(v, float) and math.isnan(v)):
        return None
    s = str(v)
    if trim:
        s = s.strip()
    if s == "" or s.lower() in {"null", "none", "nan", "na"}:
        return None

    if case_insensitive:
        s = s.lower()

    # canonicalize "001" == "1" (optional numeric)
    if numeric_canonical:
        # only if it looks like an integer/float token
        if re.fullmatch(r"[+-]?\d+", s):
            try:
                return str(int(s))
            except Exception:
                return s
        if re.fullmatch(r"[+-]?\d+(\.\d+)?", s):
            try:
                # remove trailing .0 etc
                f = float(s)
                if f.is_integer():
                    return str(int(f))
                return str(f)
            except Exception:
                return s

    return s

def stable_hash64(s: str) -> int:
    # deterministic 64-bit hash from sha1
    h = hashlib.sha1(s.encode("utf-8", errors="ignore")).digest()
    return int.from_bytes(h[:8], "little", signed=False)

@dataclass
class ColProfile:
    file: str
    table: str
    column: str
    n_rows: int
    n_nonnull: int
    n_distinct: int
    null_pct: float
    distinct_pct_nonnull: float
    avg_len: float
    looks_numeric: bool
    looks_keylike: bool
    noisy_name: bool
    reason_noisy: str

@dataclass
class ColSet:
    profile: ColProfile
    # exact set of hashes (capped)
    hashes: set
    # sketch for fast prefilter: sorted list of first K hashes
    sketch: List[int]

@dataclass
class MatchResult:
    left: ColProfile
    right: ColProfile
    inter: int
    left_distinct: int
    right_distinct: int
    contain_left: float
    contain_right: float
    jaccard: float
    strength: str
    notes: str

def infer_delimiter(path: str) -> str:
    # quick sniff
    with open(path, "r", encoding="utf-8", errors="ignore", newline="") as f:
        sample = f.read(8192)
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=[",", "\t", ";", "|"])
        return dialect.delimiter
    except Exception:
        return ","

def table_name_from_path(p: str) -> str:
    base = os.path.basename(p)
    return os.path.splitext(base)[0]

def is_probably_numeric_series(series: pd.Series) -> bool:
    # after dropping nulls
    s = series.dropna()
    if len(s) == 0:
        return False
    # if >= 90% parseable as float => numeric-ish
    ok = 0
    total = min(len(s), 2000)
    for v in s.sample(n=total, random_state=1) if len(s) > total else s:
        try:
            float(str(v).strip())
            ok += 1
        except Exception:
            pass
    return (ok / total) >= 0.90

def build_column_sets(
    csv_paths: List[str],
    case_insensitive: bool,
    numeric_canonical: bool,
    sample_rows: int,
    max_unique_per_col: int,
    sketch_k: int,
    noisy_name_regex: str
) -> Tuple[List[ColSet], List[ColProfile]]:
    colsets: List[ColSet] = []
    profiles: List[ColProfile] = []

    noisy_re = re.compile(noisy_name_regex, re.IGNORECASE)

    for path in csv_paths:
        delim = infer_delimiter(path)
        table = table_name_from_path(path)

        # read limited rows for performance
        df = pd.read_csv(path, delimiter=delim, dtype="object", nrows=sample_rows, encoding="utf-8", engine="python")
        n_rows = len(df)

        for col in df.columns:
            ser = df[col]
            n_nonnull = int(ser.notna().sum())
            null_pct = 100.0 * (1 - (n_nonnull / n_rows)) if n_rows else 100.0

            # normalize and get uniques (capped)
            uniq_hashes = set()
            lengths = []
            # sample non-nulls for speed
            s_nonnull = ser.dropna()
            if len(s_nonnull) > 0:
                s_iter = s_nonnull.sample(n=min(len(s_nonnull), sample_rows), random_state=7)
            else:
                s_iter = s_nonnull

            for v in s_iter:
                nv = normalize_value(v, case_insensitive=case_insensitive, numeric_canonical=numeric_canonical)
                if nv is None:
                    continue
                lengths.append(len(nv))
                uniq_hashes.add(stable_hash64(nv))
                if len(uniq_hashes) >= max_unique_per_col:
                    break

            n_distinct = len(uniq_hashes)
            distinct_pct_nonnull = (100.0 * n_distinct / max(n_nonnull, 1)) if n_rows else 0.0
            avg_len = (sum(lengths) / len(lengths)) if lengths else 0.0

            looks_numeric = is_probably_numeric_series(ser)
            # key-like: high distinct, low nulls, short tokens, not noisy name
            noisy_name = bool(noisy_re.search(str(col)))
            looks_keylike = (distinct_pct_nonnull >= 80.0) and (null_pct <= 10.0) and (avg_len <= 32) and (not noisy_name)

            reason_noisy = ""
            if noisy_name:
                reason_noisy = "Name pattern looks low-signal"
            elif avg_len >= 60:
                reason_noisy = "Long/free-text values"
            elif distinct_pct_nonnull <= 5.0 and n_nonnull >= 50:
                reason_noisy = "Very low distinct (enum-like)"

            prof = ColProfile(
                file=path, table=table, column=str(col),
                n_rows=n_rows, n_nonnull=n_nonnull, n_distinct=n_distinct,
                null_pct=null_pct, distinct_pct_nonnull=distinct_pct_nonnull,
                avg_len=avg_len, looks_numeric=looks_numeric,
                looks_keylike=looks_keylike, noisy_name=noisy_name,
                reason_noisy=reason_noisy
            )
            profiles.append(prof)

            sketch = sorted(uniq_hashes)[:sketch_k]
            colsets.append(ColSet(profile=prof, hashes=uniq_hashes, sketch=sketch))

    return colsets, profiles

def sketch_jaccard(a: List[int], b: List[int]) -> float:
    if not a or not b:
        return 0.0
    sa, sb = set(a), set(b)
    inter = len(sa & sb)
    union = len(sa | sb)
    return inter / union if union else 0.0

def compute_matches(
    colsets: List[ColSet],
    min_sketch_jaccard: float,
    min_intersection: int,
    min_containment: float,
    allow_same_table: bool
) -> List[MatchResult]:
    results: List[MatchResult] = []
    n = len(colsets)

    # pre-group by (numeric-ish) to reduce permutations
    groups: Dict[bool, List[int]] = {True: [], False: []}
    for i, cs in enumerate(colsets):
        groups[cs.profile.looks_numeric].append(i)

    def strength_label(pL: ColProfile, pR: ColProfile, contain_min: float, inter: int) -> Tuple[str, str]:
        # classify + notes
        if inter < min_intersection:
            return "Ignore", "Intersection too small"
        if contain_min >= 0.85 and pL.looks_keylike and pR.looks_keylike:
            return "Strong", "High containment + key-like"
        if contain_min >= 0.70 and (pL.looks_keylike or pR.looks_keylike):
            return "Potential", "Good containment; at least one side key-like"
        if contain_min >= 0.50:
            return "Weak", "Moderate overlap (often enums/codes)"
        return "Ignore", "Low overlap"

    for is_num, idxs in groups.items():
        for i_pos in range(len(idxs)):
            i = idxs[i_pos]
            A = colsets[i]
            for j_pos in range(i_pos + 1, len(idxs)):
                j = idxs[j_pos]
                B = colsets[j]

                if not allow_same_table and A.profile.table == B.profile.table:
                    continue

                # avoid obvious noisy columns (still allow if keylike)
                if (A.profile.reason_noisy and not A.profile.looks_keylike) and (B.profile.reason_noisy and not B.profile.looks_keylike):
                    continue

                # sketch prefilter
                sj = sketch_jaccard(A.sketch, B.sketch)
                if sj < min_sketch_jaccard:
                    continue

                # full intersection (on capped unique sets)
                inter = len(A.hashes & B.hashes)
                if inter < min_intersection:
                    continue

                left_dist = max(A.profile.n_distinct, 1)
                right_dist = max(B.profile.n_distinct, 1)

                contain_left = inter / left_dist
                contain_right = inter / right_dist
                contain_min = min(contain_left, contain_right)

                union = len(A.hashes | B.hashes)
                jacc = inter / union if union else 0.0

                strength, note = strength_label(A.profile, B.profile, contain_min, inter)
                if strength == "Ignore":
                    continue

                results.append(MatchResult(
                    left=A.profile, right=B.profile,
                    inter=inter, left_distinct=A.profile.n_distinct, right_distinct=A.profile.n_distinct,
                    contain_left=contain_left, contain_right=contain_right,
                    jaccard=jacc, strength=strength, notes=note
                ))

    # sort: Strong first, then containment
    order = {"Strong": 0, "Potential": 1, "Weak": 2}
    results.sort(key=lambda r: (order.get(r.strength, 9), -min(r.contain_left, r.contain_right), -r.inter))
    return results


# ---------------------------- XLSX report ----------------------------

def style_sheet_as_table(ws, table_name: str, header_row: int, last_row: int, last_col: int):
    if last_row < header_row:
        return
    ref = f"A{header_row}:{get_column_letter(last_col)}{last_row}"
    tab = Table(displayName=table_name, ref=ref)
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    ws.add_table(tab)

def autosize_columns(ws, max_width=55):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col[:5000]:  # safety
            try:
                v = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(v))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), max_width)

def apply_header_style(ws, header_row=1):
    thin = Side(style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(bold=True, color="FFFFFF")
    for cell in ws[header_row]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(vertical="center")
        cell.border = border
    ws.freeze_panes = ws[f"A{header_row+1}"]

def write_report_xlsx(out_path: str, profiles: List[ColProfile], matches: List[MatchResult], noisy_name_regex: str):
    wb = Workbook()
    wb.remove(wb.active)

    # Summary
    ws = wb.create_sheet("Summary")
    ws.append(["Metric", "Value"])
    ws.append(["Files scanned", len(set(p.file for p in profiles))])
    ws.append(["Columns scanned", len(profiles)])
    ws.append(["Join candidates", len(matches)])
    ws.append(["Strong joins", sum(1 for m in matches if m.strength == "Strong")])
    ws.append(["Potential joins", sum(1 for m in matches if m.strength == "Potential")])
    ws.append(["Weak joins", sum(1 for m in matches if m.strength == "Weak")])
    ws.append(["Noisy name regex", noisy_name_regex])
    apply_header_style(ws, 1)
    style_sheet_as_table(ws, "SummaryTable", 1, ws.max_row, ws.max_column)
    autosize_columns(ws)

    # Columns profile
    ws = wb.create_sheet("Columns_Profile")
    ws.append([
        "table", "column", "rows", "nonnull", "null_pct",
        "distinct", "distinct_pct_nonnull", "avg_len",
        "looks_numeric", "looks_keylike", "noisy_flag", "noisy_reason",
        "source_file"
    ])
    for p in profiles:
        ws.append([
            p.table, p.column, p.n_rows, p.n_nonnull, round(p.null_pct, 2),
            p.n_distinct, round(p.distinct_pct_nonnull, 2), round(p.avg_len, 1),
            p.looks_numeric, p.looks_keylike, bool(p.reason_noisy), p.reason_noisy,
            os.path.basename(p.file)
        ])
    apply_header_style(ws, 1)
    style_sheet_as_table(ws, "ColumnsProfileTable", 1, ws.max_row, ws.max_column)
    autosize_columns(ws)

    # Matches (detailed)
    ws = wb.create_sheet("Matches")
    ws.append([
        "strength",
        "left_table", "left_column",
        "right_table", "right_column",
        "intersection",
        "left_distinct", "right_distinct",
        "contain_left", "contain_right", "contain_min",
        "jaccard",
        "left_keylike", "right_keylike",
        "left_noisy_reason", "right_noisy_reason",
        "notes"
    ])
    for m in matches:
        contain_min = min(m.contain_left, m.contain_right)
        ws.append([
            m.strength,
            m.left.table, m.left.column,
            m.right.table, m.right.column,
            m.inter,
            m.left.n_distinct, m.right.n_distinct,
            round(m.contain_left, 4), round(m.contain_right, 4), round(contain_min, 4),
            round(m.jaccard, 4),
            m.left.looks_keylike, m.right.looks_keylike,
            m.left.reason_noisy, m.right.reason_noisy,
            m.notes
        ])
    apply_header_style(ws, 1)
    style_sheet_as_table(ws, "MatchesTable", 1, ws.max_row, ws.max_column)
    autosize_columns(ws)

    # Join conditions (clean)
    ws = wb.create_sheet("Join_Conditions")
    ws.append(["strength", "join_condition", "contain_min", "intersection", "notes"])
    for m in matches:
        cond = f"{m.left.table}.{m.left.column} ↔ {m.right.table}.{m.right.column}"
        contain_min = min(m.contain_left, m.contain_right)
        ws.append([m.strength, cond, round(contain_min, 4), m.inter, m.notes])
    apply_header_style(ws, 1)
    style_sheet_as_table(ws, "JoinConditionsTable", 1, ws.max_row, ws.max_column)
    autosize_columns(ws)

    # Noisy columns
    ws = wb.create_sheet("Noisy_Columns")
    ws.append(["table", "column", "why_noisy", "distinct_pct_nonnull", "avg_len", "null_pct"])
    for p in profiles:
        if p.reason_noisy and not p.looks_keylike:
            ws.append([p.table, p.column, p.reason_noisy, round(p.distinct_pct_nonnull, 2), round(p.avg_len, 1), round(p.null_pct, 2)])
    apply_header_style(ws, 1)
    style_sheet_as_table(ws, "NoisyColumnsTable", 1, ws.max_row, ws.max_column)
    autosize_columns(ws)

    wb.save(out_path)


# ---------------------------- Qt Worker ----------------------------

class Worker(QThread):
    progress = Signal(int)
    status = Signal(str)
    finished_ok = Signal(str)
    failed = Signal(str)

    def __init__(self, csv_paths, out_path, opts):
        super().__init__()
        self.csv_paths = csv_paths
        self.out_path = out_path
        self.opts = opts

    def run(self):
        try:
            self.status.emit("Reading CSVs and profiling columns…")
            self.progress.emit(5)

            colsets, profiles = build_column_sets(
                self.csv_paths,
                case_insensitive=self.opts["case_insensitive"],
                numeric_canonical=self.opts["numeric_canonical"],
                sample_rows=self.opts["sample_rows"],
                max_unique_per_col=self.opts["max_unique_per_col"],
                sketch_k=self.opts["sketch_k"],
                noisy_name_regex=self.opts["noisy_name_regex"],
            )
            self.progress.emit(45)

            self.status.emit("Finding join candidates (sketch prefilter + exact overlap)…")
            matches = compute_matches(
                colsets,
                min_sketch_jaccard=self.opts["min_sketch_jaccard"],
                min_intersection=self.opts["min_intersection"],
                min_containment=self.opts["min_containment"],
                allow_same_table=self.opts["allow_same_table"],
            )
            self.progress.emit(80)

            self.status.emit("Writing professional XLSX report…")
            write_report_xlsx(self.out_path, profiles, matches, self.opts["noisy_name_regex"])
            self.progress.emit(100)

            self.finished_ok.emit(self.out_path)
        except Exception as e:
            self.failed.emit(str(e))


# ---------------------------- Qt UI ----------------------------

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("CSV Join Finder (Value Overlap) → XLSX Report")
        self.resize(1050, 650)

        root = QWidget()
        self.setCentralWidget(root)
        layout = QVBoxLayout(root)

        # File picker area
        top = QHBoxLayout()
        layout.addLayout(top)

        left_box = QGroupBox("CSV files")
        left_layout = QVBoxLayout(left_box)
        self.list = QListWidget()
        left_layout.addWidget(self.list)

        btns = QHBoxLayout()
        self.btn_add = QPushButton("Add CSVs…")
        self.btn_clear = QPushButton("Clear")
        btns.addWidget(self.btn_add)
        btns.addWidget(self.btn_clear)
        left_layout.addLayout(btns)

        top.addWidget(left_box, 2)

        # Options
        opt_box = QGroupBox("Options (tune if Matches is empty)")
        opt_layout = QFormLayout(opt_box)

        self.sample_rows = QSpinBox()
        self.sample_rows.setRange(100, 5_000_000)
        self.sample_rows.setValue(200000)
        opt_layout.addRow("Read/sample rows per file", self.sample_rows)

        self.max_unique = QSpinBox()
        self.max_unique.setRange(500, 2_000_000)
        self.max_unique.setValue(200000)
        opt_layout.addRow("Max unique values per column (cap)", self.max_unique)

        self.sketch_k = QSpinBox()
        self.sketch_k.setRange(16, 1024)
        self.sketch_k.setValue(128)
        opt_layout.addRow("Sketch size (prefilter)", self.sketch_k)

        self.min_sketch_j = QDoubleSpinBox()
        self.min_sketch_j.setRange(0.0, 1.0)
        self.min_sketch_j.setDecimals(3)
        self.min_sketch_j.setValue(0.02)  # low to avoid empty matches
        opt_layout.addRow("Min sketch Jaccard", self.min_sketch_j)

        self.min_inter = QSpinBox()
        self.min_inter.setRange(1, 2_000_000)
        self.min_inter.setValue(10)
        opt_layout.addRow("Min intersection size", self.min_inter)

        self.min_contain = QDoubleSpinBox()
        self.min_contain.setRange(0.0, 1.0)
        self.min_contain.setDecimals(2)
        self.min_contain.setValue(0.50)
        opt_layout.addRow("Min containment (min side)", self.min_contain)

        self.case_ins = QCheckBox("Case-insensitive normalize")
        self.case_ins.setChecked(True)
        opt_layout.addRow(self.case_ins)

        self.num_canon = QCheckBox("Numeric canonicalize (001 == 1)")
        self.num_canon.setChecked(True)
        opt_layout.addRow(self.num_canon)

        self.allow_same_table = QCheckBox("Allow joins within same CSV/table")
        self.allow_same_table.setChecked(False)
        opt_layout.addRow(self.allow_same_table)

        self.noisy_regex = QLineEdit(NOISY_NAME_DEFAULT)
        opt_layout.addRow("Noisy column-name regex", self.noisy_regex)

        top.addWidget(opt_box, 3)

        # Run area
        run_bar = QHBoxLayout()
        layout.addLayout(run_bar)

        self.btn_out = QPushButton("Choose Output XLSX…")
        self.out_label = QLabel("No output file selected.")
        self.out_label.setTextInteractionFlags(Qt.TextSelectableByMouse)

        run_bar.addWidget(self.btn_out)
        run_bar.addWidget(self.out_label, 1)

        self.btn_run = QPushButton("Generate Report")
        run_bar.addWidget(self.btn_run)

        self.progress = QProgressBar()
        self.progress.setValue(0)
        layout.addWidget(self.progress)

        self.status = QLabel("")
        layout.addWidget(self.status)

        # hooks
        self.btn_add.clicked.connect(self.add_files)
        self.btn_clear.clicked.connect(self.list.clear)
        self.btn_out.clicked.connect(self.choose_out)
        self.btn_run.clicked.connect(self.run_report)

        self.out_path = None
        self.worker: Optional[Worker] = None

    def add_files(self):
        paths, _ = QFileDialog.getOpenFileNames(self, "Select CSV files", "", "CSV files (*.csv);;All files (*.*)")
        for p in paths:
            item = QListWidgetItem(p)
            self.list.addItem(item)

    def choose_out(self):
        p, _ = QFileDialog.getSaveFileName(self, "Save XLSX report as", "join_report.xlsx", "Excel Workbook (*.xlsx)")
        if p:
            if not p.lower().endswith(".xlsx"):
                p += ".xlsx"
            self.out_path = p
            self.out_label.setText(p)

    def run_report(self):
        csvs = [self.list.item(i).text() for i in range(self.list.count())]
        if not csvs:
            QMessageBox.warning(self, "Missing input", "Please add at least one CSV file.")
            return
        if not self.out_path:
            QMessageBox.warning(self, "Missing output", "Please choose an output XLSX file.")
            return

        opts = {
            "sample_rows": int(self.sample_rows.value()),
            "max_unique_per_col": int(self.max_unique.value()),
            "sketch_k": int(self.sketch_k.value()),
            "min_sketch_jaccard": float(self.min_sketch_j.value()),
            "min_intersection": int(self.min_inter.value()),
            "min_containment": float(self.min_contain.value()),
            "case_insensitive": bool(self.case_ins.isChecked()),
            "numeric_canonical": bool(self.num_canon.isChecked()),
            "allow_same_table": bool(self.allow_same_table.isChecked()),
            "noisy_name_regex": self.noisy_regex.text().strip() or NOISY_NAME_DEFAULT
        }

        self.progress.setValue(0)
        self.status.setText("Starting…")
        self.btn_run.setEnabled(False)

        self.worker = Worker(csvs, self.out_path, opts)
        self.worker.progress.connect(self.progress.setValue)
        self.worker.status.connect(self.status.setText)
        self.worker.finished_ok.connect(self.on_done)
        self.worker.failed.connect(self.on_fail)
        self.worker.start()

    def on_done(self, out_path: str):
        self.btn_run.setEnabled(True)
        self.status.setText(f"Done. Report generated: {out_path}")
        QMessageBox.information(self, "Report generated", f"Report saved:\n{out_path}")

    def on_fail(self, err: str):
        self.btn_run.setEnabled(True)
        self.status.setText("Failed.")
        QMessageBox.critical(self, "Error", err)


def main():
    app = QApplication(sys.argv)
    w = MainWindow()
    w.show()
    sys.exit(app.exec())

if __name__ == "__main__":
    main()
