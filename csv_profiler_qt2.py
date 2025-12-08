import os
import sys
import csv
import subprocess
import io

from PyQt6.QtCore import Qt, QThread, pyqtSignal
from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QLineEdit, QLabel, QFileDialog, QTableWidget,
    QTableWidgetItem, QHeaderView, QProgressBar, QMessageBox,
    QPlainTextEdit, QCheckBox
)


# ------------------------------
# Worker thread for profiling
# ------------------------------
class ProfilingWorker(QThread):
    progress = pyqtSignal(int, int, str)                          # current, total, message
    fileUpdated = pyqtSignal(int, str, int, int, str, str)        # index, status, row_count, col_count, model_role, hierarchy_hints
    finished = pyqtSignal(str)                                    # report_path
    error = pyqtSignal(str)                                       # fatal error message

    def __init__(
        self,
        files,
        root_path,
        report_name="csv_profiling_report.xlsx",
        generate_unique_files=True,
        unique_folder_name="unique_values",
        output_root_path=None,
        parent=None,
    ):
        """
        :param files: list of dicts: {"subfolder": str, "name": str, "path": str}
        :param root_path: selected root folder path (where CSVs live)
        :param output_root_path: where to write outputs (one level up)
        """
        super().__init__(parent)
        self.files = files
        self.root_path = root_path
        self.report_name = report_name
        self.generate_unique_files = generate_unique_files
        self.unique_folder_name = unique_folder_name
        self.output_root_path = output_root_path or root_path  # fallback to root if None

    def run(self):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            self.error.emit(
                "openpyxl is not installed. Install it with: pip install openpyxl"
            )
            return

        total = len(self.files)
        if total == 0:
            self.error.emit("No CSV files found to profile.")
            return

        wb = Workbook()
        ws_prof = wb.active
        ws_prof.title = "CSV Profiling"

        header_written = False
        summary_rows = []   # For Summary sheet (including auto model role + hierarchy hints)
        sample_rows = []    # For Sample Value sheet

        # NOTE: outputs live under self.output_root_path
        unique_root = None
        if self.generate_unique_files:
            try:
                unique_root = os.path.join(self.output_root_path, self.unique_folder_name)
                os.makedirs(unique_root, exist_ok=True)
            except Exception as e:
                # Log and continue without unique files
                self.progress.emit(0, total, f"ERROR creating unique_values folder: {e}")
                unique_root = None

        for idx, f in enumerate(self.files):
            subfolder = f.get("subfolder", "")
            filename = f.get("name", "")
            full_path = f.get("path", "")

            # REMOVE .csv FROM DATASET NAME FOR EXCEL
            dataset_name, _ = os.path.splitext(filename)

            self.progress.emit(idx + 1, total, f"Profiling: {filename}")

            # Run csvstat --csv
            try:
                result = subprocess.run(
                    ["csvstat", "--csv", full_path],
                    capture_output=True,
                    text=True
                )
                if result.returncode != 0 or not result.stdout.strip():
                    msg = result.stderr.strip() or "csvstat returned an error."
                    self.fileUpdated.emit(idx, f"Error: {msg}", 0, 0, "", "")
                    self.progress.emit(idx + 1, total, f"ERROR csvstat for {filename}: {msg}")
                    # continue to next file
                    continue
            except Exception as e:
                self.fileUpdated.emit(idx, f"csvstat failed: {e}", 0, 0, "", "")
                self.progress.emit(idx + 1, total, f"ERROR running csvstat for {filename}: {e}")
                continue

            # Parse csvstat output
            try:
                reader = csv.reader(io.StringIO(result.stdout))
                try:
                    header = next(reader)
                except StopIteration:
                    self.fileUpdated.emit(idx, "No stats output", 0, 0, "", "")
                    self.progress.emit(idx + 1, total, f"ERROR: No csvstat stats for {filename}")
                    continue

                if not header_written:
                    # Excel header: Source (subfolder), Dataset Name (file), + csvstat columns
                    ws_prof.append(["Source", "Dataset Name", *header])
                    header_written = True

                # Find 'type' column index in csvstat output to build data type counts
                type_idx = -1
                for i, col_name in enumerate(header):
                    if col_name.strip().lower() == "type":
                        type_idx = i
                        break
                type_counts = {}

                # Quick file-level stats: row & column count
                row_count, col_count = self._quick_stats(full_path)

                for row in reader:
                    # Count types for Summary
                    try:
                        if type_idx != -1 and type_idx < len(row):
                            tval = row[type_idx].strip()
                            if tval:
                                type_counts[tval] = type_counts.get(tval, 0) + 1
                    except Exception:
                        # ignore bad type rows
                        pass

                    # USE dataset_name WITHOUT .csv
                    try:
                        ws_prof.append([subfolder, dataset_name, *row])
                    except Exception as e:
                        # Log but keep going with other rows
                        self.progress.emit(
                            idx + 1,
                            total,
                            f"ERROR writing profiling row for {filename}: {e}"
                        )

            except Exception as e:
                self.fileUpdated.emit(idx, f"Parse error: {e}", 0, 0, "", "")
                self.progress.emit(idx + 1, total, f"ERROR parsing csvstat output for {filename}: {e}")
                # move on but do NOT add to summary_rows
                continue

            # Collect samples (always, independent of unique-values files)
            col_samples = []
            try:
                self.progress.emit(idx + 1, total, f"Collecting samples: {filename}")
                col_samples = self._collect_samples(full_path, subfolder, dataset_name)
                sample_rows.extend(col_samples)
            except Exception as e:
                self.progress.emit(idx + 1, total, f"ERROR collecting samples for {filename}: {e}")

            # Infer model role (Fact/Dim/etc.) using heuristics
            try:
                model_role = self._infer_model_role(
                    col_samples=col_samples,
                    row_count=row_count,
                    dataset_name=dataset_name,
                )
            except Exception:
                model_role = ""

            # Detect hierarchies from column names
            try:
                hierarchy_hints_list = self._detect_hierarchies(
                    [c["col_name"] for c in col_samples]
                )
                hierarchy_hints = "; ".join(hierarchy_hints_list)
            except Exception:
                hierarchy_hints = ""

            # Capture info for summary sheet (add model_role + hierarchy hints here)
            try:
                summary_rows.append(
                    {
                        "source": subfolder,
                        "dataset": dataset_name,
                        "row_count": row_count,
                        "col_count": col_count,
                        "type_counts": type_counts,
                        "model_role": model_role,
                        "hierarchy_hints": hierarchy_hints,
                    }
                )
            except Exception as e:
                self.progress.emit(0, len(summary_rows), f"ERROR adding to Summary rows: {e}")

            # Generate per-column unique values files (optional)
            if self.generate_unique_files and unique_root is not None:
                try:
                    self.progress.emit(idx + 1, total, f"Unique values: {filename}")
                    self._generate_unique_values_files(
                        full_path=full_path,
                        unique_root=unique_root,
                        subfolder=subfolder,
                        dataset_name=dataset_name,
                    )
                except Exception as e:
                    # Log and continue
                    self.progress.emit(idx + 1, total, f"ERROR unique values for {filename}: {e}")

            # Now that everything for this file is ready, update UI row once
            self.fileUpdated.emit(idx, "Done", row_count, col_count, model_role, hierarchy_hints)

        if not header_written:
            self.error.emit("No profiling data generated (csvstat may have failed for all files).")
            return

        # ------------------------------
        # CREATE SUMMARY SHEET (FIRST TAB)
        # ------------------------------
        ws_summary = wb.create_sheet(title="Summary", index=0)
        ws_summary.append([
            "Source",
            "Dataset Name",
            "Model Role (auto)",
            "Hierarchy Hints",
            "Row Count",
            "Column Count",
            "Data Type Counts",
        ])

        for entry in summary_rows:
            try:
                type_counts = entry.get("type_counts") or {}
                if type_counts:
                    parts = [
                        f"{t}={c}"
                        for t, c in sorted(type_counts.items(), key=lambda x: x[0].lower())
                    ]
                    type_counts_str = "; ".join(parts)
                else:
                    type_counts_str = ""

                ws_summary.append([
                    entry.get("source", ""),
                    entry.get("dataset", ""),
                    entry.get("model_role", ""),
                    entry.get("hierarchy_hints", ""),
                    entry.get("row_count", 0),
                    entry.get("col_count", 0),
                    type_counts_str,
                ])
            except Exception as e:
                # Log and continue
                self.progress.emit(0, len(summary_rows), f"ERROR writing Summary row: {e}")

        # ------------------------------
        # CREATE SAMPLE VALUE SHEET (THIRD TAB)
        # ------------------------------
        ws_sample = wb.create_sheet(title="Sample Value")
        ws_sample.append([
            "Source",
            "Dataset Name",
            "column_name",
            "data_type",
            "sample_value_1",
            "sample_value_2",
            "sample_value_3",
            "sample_value_4",
            "sample_value_5",
        ])

        for row in sample_rows:
            try:
                samples = row.get("samples", [])
                s1 = samples[0] if len(samples) > 0 else ""
                s2 = samples[1] if len(samples) > 1 else ""
                s3 = samples[2] if len(samples) > 2 else ""
                s4 = samples[3] if len(samples) > 3 else ""
                s5 = samples[4] if len(samples) > 4 else ""
                ws_sample.append([
                    row.get("source", ""),
                    row.get("dataset", ""),
                    row.get("col_name", ""),
                    row.get("data_type", ""),
                    s1,
                    s2,
                    s3,
                    s4,
                    s5,
                ])
            except Exception as e:
                self.progress.emit(0, len(sample_rows), f"ERROR writing Sample Value row: {e}")

        # ------------------------------
        # PROFESSIONAL FORMATTING (B/W) FOR ALL SHEETS
        # ------------------------------
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        def style_sheet(ws, col_widths=None):
            try:
                max_row = ws.max_row
                max_col = ws.max_column

                # Freeze header row and add autofilter
                ws.freeze_panes = "A2"
                ws.auto_filter.ref = ws.dimensions

                thin_border = Border(
                    left=Side(style="thin", color="000000"),
                    right=Side(style="thin", color="000000"),
                    top=Side(style="thin", color="000000"),
                    bottom=Side(style="thin", color="000000"),
                )

                # Header style: black background, white bold text
                for cell in ws[1]:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill("solid", fgColor="000000")
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    cell.border = thin_border

                # Body rows: zebra striping (white / light gray) + borders
                for row in ws.iter_rows(min_row=2, max_row=max_row, max_col=max_col):
                    is_alt = (row[0].row % 2 == 0)
                    fill_color = "FFFFFF" if not is_alt else "F7F7F7"
                    for cell in row:
                        cell.fill = PatternFill("solid", fgColor=fill_color)
                        cell.border = thin_border
                        cell.alignment = Alignment(vertical="top")

                # Column widths
                if col_widths:
                    for col_letter, width in col_widths.items():
                        ws.column_dimensions[col_letter].width = width
            except Exception:
                # Styling failure is non-fatal
                pass

        # Summary widths
        style_sheet(
            ws_summary,
            col_widths={
                "A": 30,  # Source
                "B": 40,  # Dataset Name
                "C": 25,  # Model Role
                "D": 50,  # Hierarchy Hints
                "E": 15,  # Row Count
                "F": 15,  # Column Count
                "G": 50,  # Data Type Counts
            },
        )

        # Sample Value widths (now has 9 columns: A..I)
        style_sheet(
            ws_sample,
            col_widths={
                "A": 30,  # Source
                "B": 40,  # Dataset
                "C": 40,  # column_name
                "D": 15,  # data_type
                "E": 25,  # sample 1
                "F": 25,  # sample 2
                "G": 25,  # sample 3
                "H": 25,  # sample 4
                "I": 25,  # sample 5
            },
        )

        # Profiling widths (same as before)
        style_sheet(
            ws_prof,
            col_widths={
                "A": 30,  # Source
                "B": 40,  # Dataset Name
                # rest default 18 below
            },
        )
        # For profiling, set default width for other columns
        try:
            for col_cells in ws_prof.columns:
                col_letter = col_cells[0].column_letter
                if col_letter in ("A", "B"):
                    continue
                ws_prof.column_dimensions[col_letter].width = 18
        except Exception:
            pass

        # ------------------------------
        # SAVE WORKBOOK ONE LEVEL UP
        # ------------------------------
        try:
            report_path = os.path.join(self.output_root_path, self.report_name)
            wb.save(report_path)
        except Exception as e:
            self.error.emit(f"Error saving Excel report: {e}")
            return

        self.finished.emit(report_path)

    # --------- helpers ---------
    def _quick_stats(self, full_path):
        """Rudimentary row & column count using Python's csv module."""
        row_count = 0
        col_count = 0
        try:
            with open(full_path, "r", newline="", encoding="utf-8", errors="ignore") as f:
                reader = csv.reader(f)
                for i, row in enumerate(reader):
                    if i == 0:
                        col_count = len(row)
                    else:
                        row_count += 1
        except Exception:
            # non-fatal, leave row_count/col_count as 0
            pass
        return row_count, col_count

    def _sanitize_filename_part(self, text: str) -> str:
        # Only allow letters, digits, -, _ ; everything else becomes _
        if text is None:
            return ""
        return "".join(ch if (ch.isalnum() or ch in ("-", "_")) else "_" for ch in text)[:150]

    def _infer_type(self, values):
        """Simple type inference based on unique values."""
        try:
            non_empty = [v for v in values if v not in ("", None)]
            if not non_empty:
                return "text"

            all_int = True
            all_float = True

            for v in non_empty:
                try:
                    int(v)
                except Exception:
                    all_int = False
                try:
                    float(v)
                except Exception:
                    all_float = False

            if all_int:
                return "int"
            if all_float:
                return "float"
            return "text"
        except Exception:
            return "text"

    def _collect_samples(self, full_path, subfolder, dataset_name):
        """
        Collect up to 5 distinct sample values per column + inferred data type.
        Returns list of dicts:
        {
          "source": subfolder,
          "dataset": dataset_name,
          "col_name": column_name,
          "data_type": type_str,
          "samples": [v1..v5],
          "unique_count": int
        }
        """
        rows = []
        try:
            with open(full_path, "r", newline="", encoding="utf-8", errors="ignore") as f:
                reader = csv.reader(f)
                try:
                    header = next(reader)
                except StopIteration:
                    return rows

                num_cols = len(header)
                col_names = [h if h is not None else f"col_{i+1}" for i, h in enumerate(header)]
                unique_sets = [set() for _ in range(num_cols)]

                row_counter = 0
                for row in reader:
                    row_counter += 1
                    if len(row) < num_cols:
                        row = list(row) + [""] * (num_cols - len(row))
                    elif len(row) > num_cols:
                        row = row[:num_cols]

                    for idx in range(num_cols):
                        val = row[idx]
                        if val not in ("", None):
                            unique_sets[idx].add(val)

                for idx in range(num_cols):
                    col_name = col_names[idx] or f"col_{idx+1}"
                    values = list(unique_sets[idx])
                    try:
                        values_sorted = sorted(values, key=lambda x: (x is None, str(x)))
                    except Exception:
                        values_sorted = values
                    samples = values_sorted[:5]
                    data_type = self._infer_type(values)

                    rows.append(
                        {
                            "source": subfolder,
                            "dataset": dataset_name,
                            "col_name": col_name,
                            "data_type": data_type,
                            "samples": samples,
                            "unique_count": len(unique_sets[idx]),
                            # row_counter is approximate row_count for this file
                        }
                    )
        except Exception:
            # non-fatal, just return what we have
            return rows
        return rows

    def _generate_unique_values_files(self, full_path, unique_root, subfolder, dataset_name):
        """
        For each column in the CSV:
        - Collect unique values
        - Infer simple data type
        - Write to: unique_root / subfolder / <dataset>_<colnum>_<colname>_<datatype>.txt
        """
        safe_dataset = self._sanitize_filename_part(dataset_name)

        target_dir = unique_root
        if subfolder and subfolder != ".":
            target_dir = os.path.join(unique_root, subfolder)
        try:
            os.makedirs(target_dir, exist_ok=True)
        except Exception:
            # non-fatal
            return

        try:
            with open(full_path, "r", newline="", encoding="utf-8", errors="ignore") as f:
                reader = csv.reader(f)
                try:
                    header = next(reader)
                except StopIteration:
                    return

                num_cols = len(header)
                col_names = [h if h is not None else f"col_{i+1}" for i, h in enumerate(header)]
                unique_sets = [set() for _ in range(num_cols)]

                for row in reader:
                    # Pad or trim to num_cols
                    if len(row) < num_cols:
                        row = list(row) + [""] * (num_cols - len(row))
                    elif len(row) > num_cols:
                        row = row[:num_cols]

                    for idx in range(num_cols):
                        val = row[idx]
                        unique_sets[idx].add(val)
        except Exception:
            # non-fatal
            return

        # Now write one file per column
        for idx in range(num_cols):
            try:
                col_num = idx + 1  # 1-based
                col_name = col_names[idx] or f"col_{col_num}"
                values = list(unique_sets[idx])

                data_type = self._infer_type(values)

                safe_col_name = self._sanitize_filename_part(col_name)
                filename = f"{safe_dataset}_{col_num}_{safe_col_name}_{data_type}.txt"
                out_path = os.path.join(target_dir, filename)

                # Sort values for readability
                try:
                    values_sorted = sorted(values, key=lambda x: (x is None, str(x)))
                except Exception:
                    values_sorted = values

                with open(out_path, "w", encoding="utf-8", newline="\n") as out_f:
                    for v in values_sorted:
                        out_f.write("" if v is None else str(v))
                        out_f.write("\n")
            except Exception:
                # continue with next column
                continue

    # --------- fact/dim + hierarchy heuristic ---------
    def _infer_model_role(self, col_samples, row_count, dataset_name):
        """
        Heuristic guess of model role for a CSV using Fact vs Dimension rules:

        FACT clues:
        - multiple numeric measure-like columns
        - date/timestamp columns
        - many foreign-key-ish IDs
        - large row counts (10k+ / 100k+ / 1M+)
        - event/line-level names (order_line, transaction, usage, log, etc.)
        - minimal descriptive columns

        DIMENSION clues:
        - many descriptive text attributes (name, type, category, region, etc.)
        - small/medium row count (< 50k)
        - one main business key with near-unique values
        - few or no measures
        - stable-looking attributes (addresses, names, etc.)

        Returns e.g.:
        - "Fact (auto)"
        - "Likely Fact (auto)"
        - "Dimension (auto)"
        - "Likely Dimension/Ref (auto)"
        - "Mixed/Bridge? (auto)"
        """
        if not col_samples:
            return "Unknown (auto)"

        col_names = [c["col_name"] for c in col_samples]
        dtype_map = {c["col_name"]: c.get("data_type", "text") for c in col_samples}
        uniq_map = {c["col_name"]: c.get("unique_count", 0) for c in col_samples}

        total_cols = max(len(col_names), 1)

        # --- 1. Numeric measures ---
        numeric_cols = [c for c in col_names if dtype_map.get(c) in ("int", "float")]

        measure_keywords = [
            "amount", "amt", "qty", "quantity", "price", "tax", "discount",
            "unit", "units", "balance", "balances", "cost", "rate",
            "total", "revenue", "sales", "profit", "margin",
            "count", "cnt", "duration", "hours", "mins", "minutes",
            "score", "weight", "volume"
        ]

        numeric_measure_cols = [
            c for c in numeric_cols
            if any(kw in c.lower() for kw in measure_keywords)
            and not c.lower().endswith("_id")
            and "key" not in c.lower()
        ]

        # --- 2. ID / foreign-key style columns ---
        fk_suffix_keywords = ["_id", "_key", "_code", "_number", "id_", "key_", "code_"]
        fk_entity_keywords = [
            "customer", "client", "member", "account", "product", "item", "sku",
            "vendor", "supplier", "store", "location", "branch", "employee",
            "user", "agent", "policy", "invoice", "order", "contract", "claim",
            "ticket", "shipment", "warehouse", "category"
        ]

        fk_cols = []
        for c in col_names:
            lc = c.lower()
            if any(kw in lc for kw in fk_suffix_keywords) or any(ent in lc for ent in fk_entity_keywords):
                fk_cols.append(c)
        fk_cols = list(set(fk_cols))

        # --- 3. Date / time columns ---
        date_keywords = [
            "date", "_dt", "_at", "time", "timestamp", "created", "updated",
            "posted", "effective", "start", "end", "from", "to", "period",
            "month", "year", "week", "day"
        ]
        date_cols = [c for c in col_names if any(kw in c.lower() for kw in date_keywords)]

        # --- 4. Descriptive columns ---
        desc_keywords = [
            "name", "description", "desc", "address", "phone", "email",
            "city", "state", "country", "zip", "postal", "comment", "comments",
            "remark", "remarks", "note", "notes", "type", "category",
            "segment", "brand", "region", "status"
        ]
        desc_cols = [c for c in col_names if any(kw in c.lower() for kw in desc_keywords)]

        # --- 5. Event / line-level naming ---
        event_keywords = [
            "line", "detail", "item", "txn", "transaction", "usage", "log",
            "event", "movement", "entry", "posting", "click",
            "impression", "shipment", "history", "fact"
        ]
        name_lc = dataset_name.lower()
        name_event_flag = any(kw in name_lc for kw in event_keywords)
        col_event_flag = any(any(kw in c.lower() for kw in event_keywords) for c in col_names)

        # --- 6. Candidate dimension primary keys ---
        candidate_dim_keys = []
        if row_count > 0:
            for c in col_names:
                lc = c.lower()
                # looks like business key
                if (
                    any(ent in lc for ent in fk_entity_keywords)
                    or lc.endswith("_id")
                    or lc.endswith("_key")
                    or lc.endswith("_code")
                ):
                    uniq = uniq_map.get(c, 0)
                    if uniq == 0:
                        continue
                    ratio = uniq / float(row_count)
                    # near-unique
                    if ratio >= 0.7:
                        candidate_dim_keys.append(c)

        # --- Fact & Dimension scoring ---
        fact_score = 0.0
        dim_score = 0.0

        # FACT: numeric measures
        if len(numeric_measure_cols) >= 2:
            fact_score += 2.0
        elif len(numeric_measure_cols) == 1:
            fact_score += 1.0

        # FACT: date columns
        if date_cols:
            fact_score += 1.0

        # FACT: multiple foreign keys (linking dimensions)
        if len(fk_cols) >= 3:
            fact_score += 1.5
        elif len(fk_cols) >= 2:
            fact_score += 1.0

        # FACT: row count
        if row_count >= 1_000_000:
            fact_score += 2.5
        elif row_count >= 100_000:
            fact_score += 2.0
        elif row_count >= 10_000:
            fact_score += 1.5
        elif row_count >= 1_000:
            fact_score += 1.0

        # FACT: event-ish naming
        if name_event_flag:
            fact_score += 1.0
        if col_event_flag:
            fact_score += 1.0

        # FACT penalty: heavy descriptive text suggests dimension instead
        if len(desc_cols) >= max(3, int(total_cols * 0.3)):
            fact_score -= 1.0

        # DIM: descriptive-heavy
        if len(desc_cols) >= max(1, int(total_cols * 0.2)):
            dim_score += 1.5
        if len(desc_cols) >= max(3, int(total_cols * 0.4)):
            dim_score += 0.5

        # DIM: small/medium row counts
        if row_count > 0:
            if row_count < 1_000:
                dim_score += 2.0
            elif row_count < 10_000:
                dim_score += 1.5
            elif row_count < 50_000:
                dim_score += 1.0

        # DIM: few measures
        if len(numeric_measure_cols) == 0 and len(numeric_cols) <= 3:
            dim_score += 1.0

        # DIM: candidate primary key(s)
        if len(candidate_dim_keys) == 1:
            dim_score += 2.0
        elif len(candidate_dim_keys) > 1:
            dim_score += 1.5

        # Quick override for classic small descriptive lookup/status dims
        if (
            row_count < 5_000
            and len(numeric_measure_cols) == 0
            and len(desc_cols) >= 1
            and len(candidate_dim_keys) <= 2
        ):
            return "Dimension (auto)"

        # Now compare scores
        if fact_score >= dim_score + 1.0:
            # Strong fact
            if fact_score >= 4.5:
                return "Fact (auto)"
            else:
                return "Likely Fact (auto)"
        elif dim_score >= fact_score + 1.0:
            # Strong dimension
            if dim_score >= 4.0:
                return "Dimension (auto)"
            else:
                return "Likely Dimension/Ref (auto)"
        else:
            # Ambiguous / hybrid
            return "Mixed/Bridge? (auto)"

    def _detect_hierarchies(self, col_names):
        """
        Heuristically detect hierarchies using common patterns:
        - Geo: country > state > city
        - Product: category > subcategory > item
        - Org: company > division > department
        - Time: year > quarter > month > day
        - Sales: region > territory > salesperson
        Returns a list of short human-readable hints.
        """
        hints = []
        cols_lc = [c.lower() for c in col_names]

        def has_any(tokens):
            return any(t in col for col in cols_lc for t in tokens)

        def find_exact(token):
            for c in cols_lc:
                if c == token:
                    return True
            return False

        # Geo hierarchy
        if has_any(["country"]) and has_any(["state", "province", "region"]) and has_any(["city", "town"]):
            hints.append("Geo: country > state/region > city")

        # Product hierarchy
        if has_any(["category"]) and has_any(["sub_category", "subcategory"]) and has_any(["product", "item", "sku"]):
            hints.append("Product: category > subcategory > product")
        elif has_any(["category"]) and has_any(["product", "item", "sku"]):
            hints.append("Product: category > product")

        # Time hierarchy
        if has_any(["year"]) and has_any(["quarter"]) and has_any(["month"]):
            hints.append("Time: year > quarter > month")
        elif has_any(["year"]) and has_any(["month"]):
            hints.append("Time: year > month")
        if has_any(["month"]) and has_any(["day"]) and not any("date" in c for c in cols_lc):
            hints.append("Time: month > day")

        # Org hierarchy
        if has_any(["company"]) and has_any(["division"]) and has_any(["department"]):
            hints.append("Org: company > division > department")
        elif has_any(["department"]) and has_any(["team"]):
            hints.append("Org: department > team")

        # Sales hierarchy
        if has_any(["region"]) and has_any(["territory"]) and has_any(["salesperson", "sales_person", "rep"]):
            hints.append("Sales: region > territory > salesperson")

        # Brand hierarchy
        if has_any(["brand"]) and has_any(["product_line", "line"]) and has_any(["product", "sku", "item"]):
            hints.append("Brand: brand > product line > product")

        # Region hierarchy variant
        if has_any(["region"]) and has_any(["subregion", "sub_region"]):
            hints.append("Geo: region > subregion")

        # De-duplicate
        seen = set()
        deduped = []
        for h in hints:
            if h not in seen:
                seen.add(h)
                deduped.append(h)

        return deduped


# ------------------------------
# Main Window
# ------------------------------
class CsvProfilerWindow(QMainWindow):
    def __init__(self):
        super().__init__()

        self.setWindowTitle("CSV Profiler (csvkit + Excel) - Fact/Dim Modeling Helper")
        self.resize(1200, 720)

        self.worker = None
        self.files = []
        self.report_path = None

        self._build_ui()

    def _build_ui(self):
        central = QWidget()
        self.setCentralWidget(central)

        main_layout = QVBoxLayout()
        central.setLayout(main_layout)

        # --- Folder selection row ---
        folder_layout = QHBoxLayout()

        self.folder_edit = QLineEdit()
        self.folder_edit.setPlaceholderText("Select root folder containing CSV subfolders...")
        self.folder_edit.setReadOnly(True)

        browse_btn = QPushButton("Browse...")
        browse_btn.clicked.connect(self.browse_folder)

        open_root_btn = QPushButton("Open Root Folder")
        open_root_btn.clicked.connect(self.open_root_folder)

        folder_layout.addWidget(QLabel("Root Folder:"))
        folder_layout.addWidget(self.folder_edit)
        folder_layout.addWidget(browse_btn)
        folder_layout.addWidget(open_root_btn)

        main_layout.addLayout(folder_layout)

        # --- Unique values checkbox ---
        self.unique_checkbox = QCheckBox(
            "Generate per-column unique values text files (unique_values folder, one level up)"
        )
        self.unique_checkbox.setChecked(True)
        main_layout.addWidget(self.unique_checkbox)

        # --- Buttons row ---
        buttons_layout = QHBoxLayout()

        self.scan_btn = QPushButton("Scan & Profile")
        self.scan_btn.clicked.connect(self.start_scan_and_profile)

        self.open_report_btn = QPushButton("Open Excel Report")
        self.open_report_btn.setEnabled(False)
        self.open_report_btn.clicked.connect(self.open_report)

        clear_btn = QPushButton("Clear Table")
        clear_btn.clicked.connect(self.clear_table)

        buttons_layout.addWidget(self.scan_btn)
        buttons_layout.addWidget(self.open_report_btn)
        buttons_layout.addWidget(clear_btn)
        buttons_layout.addStretch()

        main_layout.addLayout(buttons_layout)

        # --- Progress bar and status label ---
        progress_layout = QHBoxLayout()
        self.progress_bar = QProgressBar()
        self.progress_bar.setMinimum(0)
        self.progress_bar.setMaximum(100)
        self.progress_bar.setValue(0)

        self.status_label = QLabel("Ready.")
        self.status_label.setMinimumWidth(300)

        progress_layout.addWidget(self.progress_bar)
        progress_layout.addWidget(self.status_label)

        main_layout.addLayout(progress_layout)

        # --- Table of datasets ---
        # Added 8th column: Hierarchy Hints
        self.table = QTableWidget(0, 8)
        self.table.setHorizontalHeaderLabels([
            "Source (Subfolder)",
            "Dataset Name",
            "Full Path",
            "Status",
            "Model Role (Fact/Dim/Bridge/etc.)",
            "Row Count (approx)",
            "Column Count",
            "Hierarchy Hints"
        ])
        header = self.table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(5, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(6, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(7, QHeaderView.ResizeMode.Stretch)

        # Double-click to open CSV in default app (except when editing Model Role / Hierarchy)
        self.table.cellDoubleClicked.connect(self.on_cell_double_clicked)

        main_layout.addWidget(self.table)

        # --- Fact/Dim modeling hint ---
        hint_label = QLabel(
            "💡 Summary + Sample Value (with data_type & unique_count) + unique_values + auto Model Role\n"
            "   + Hierarchy Hints give you fast visibility into grains, candidate Facts/Dimensions,\n"
            "   and natural hierarchies (geo, product, time, org) without opening every file."
        )
        hint_label.setWordWrap(True)
        main_layout.addWidget(hint_label)

        # --- Log window ---
        self.log_box = QPlainTextEdit()
        self.log_box.setReadOnly(True)
        self.log_box.setPlaceholderText("Log output from csvstat and processing will appear here...")
        main_layout.addWidget(self.log_box)

    # ------------------------------
    # UI slots and helpers
    # ------------------------------
    def log(self, text: str):
        self.log_box.appendPlainText(text)

    def browse_folder(self):
        try:
            folder = QFileDialog.getExistingDirectory(self, "Select Root Folder", "")
            if folder:
                self.folder_edit.setText(folder)
                self.log(f"Selected root folder: {folder}")
        except Exception as e:
            self.log(f"ERROR selecting folder: {e}")

    def open_root_folder(self):
        root = self.folder_edit.text().strip()
        if not root or not os.path.isdir(root):
            QMessageBox.warning(self, "No Folder", "Please select a valid root folder first.")
            return
        try:
            os.startfile(root)
        except Exception as e:
            self.log(f"ERROR opening root folder: {e}")
            QMessageBox.critical(self, "Error", f"Cannot open folder: {e}")

    def clear_table(self):
        try:
            self.table.setRowCount(0)
            self.files = []
            self.report_path = None
            self.open_report_btn.setEnabled(False)
            self.progress_bar.setValue(0)
            self.status_label.setText("Cleared.")
            self.log("Cleared table and state.")
        except Exception as e:
            self.log(f"ERROR clearing table: {e}")

    def start_scan_and_profile(self):
        try:
            root = self.folder_edit.text().strip()
            if not root or not os.path.isdir(root):
                QMessageBox.warning(self, "Invalid Folder", "Please select a valid root folder.")
                return

            # Scan for CSV files
            self.files = self._scan_for_csv(root)
            self._populate_table()

            if not self.files:
                QMessageBox.information(self, "No CSV Files", "No .csv files found under this folder.")
                self.status_label.setText("No CSV files found.")
                return

            # Outputs one level up
            # If root has no parent (e.g., C:\), we fallback to root itself
            parent = os.path.dirname(root.rstrip("\\/")) or root

            report_name = "csv_profiling_report.xlsx"
            report_path = os.path.join(parent, report_name)

            # Ask about overwriting if report already exists (one level up)
            if os.path.exists(report_path):
                resp = QMessageBox.question(
                    self,
                    "Overwrite Existing Report?",
                    f"Report '{report_name}' already exists at:\n{report_path}\n\nDo you want to overwrite it?",
                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
                )
                if resp != QMessageBox.StandardButton.Yes:
                    self.log("User chose not to overwrite existing report.")
                    return

            self.report_path = None
            self.open_report_btn.setEnabled(False)

            generate_uniques = self.unique_checkbox.isChecked()

            # Start worker thread
            self.worker = ProfilingWorker(
                self.files,
                root_path=root,
                report_name=report_name,
                generate_unique_files=generate_uniques,
                output_root_path=parent,
            )
            self.worker.progress.connect(self.on_worker_progress)
            self.worker.fileUpdated.connect(self.on_worker_file_updated)
            self.worker.finished.connect(self.on_worker_finished)
            self.worker.error.connect(self.on_worker_error)

            self.scan_btn.setEnabled(False)
            self.status_label.setText("Running csvstat on all files...")
            self.progress_bar.setValue(0)
            self.log(f"Starting profiling in background thread... Outputs will go to: {parent}")

            self.worker.start()
        except Exception as e:
            self.log(f"ERROR starting scan/profile: {e}")
            QMessageBox.critical(self, "Error", f"Unexpected error: {e}")

    def _scan_for_csv(self, root):
        files = []
        try:
            for dirpath, dirnames, filenames in os.walk(root):
                for fn in filenames:
                    if fn.lower().endswith(".csv"):
                        full_path = os.path.join(dirpath, fn)
                        rel = os.path.relpath(full_path, root)
                        subfolder = os.path.dirname(rel)
                        if not subfolder:
                            subfolder = "."  # directly under root
                        files.append({
                            "subfolder": subfolder,
                            "name": fn,
                            "path": full_path
                        })
            files.sort(key=lambda f: (f["subfolder"].lower(), f["name"].lower()))
            self.log(f"Found {len(files)} CSV file(s) under {root}.")
        except Exception as e:
            self.log(f"ERROR scanning for CSV files: {e}")
        return files

    def _populate_table(self):
        try:
            self.table.setRowCount(0)
            for f in self.files:
                row = self.table.rowCount()
                self.table.insertRow(row)

                # Source (Subfolder)
                item_source = QTableWidgetItem(f.get("subfolder", ""))
                item_source.setFlags(item_source.flags() & ~Qt.ItemFlag.ItemIsEditable)
                self.table.setItem(row, 0, item_source)

                # Dataset Name (UI keeps .csv, Excel strips it)
                item_name = QTableWidgetItem(f.get("name", ""))
                item_name.setFlags(item_name.flags() & ~Qt.ItemFlag.ItemIsEditable)
                self.table.setItem(row, 1, item_name)

                # Full Path
                item_path = QTableWidgetItem(f.get("path", ""))
                item_path.setFlags(item_path.flags() & ~Qt.ItemFlag.ItemIsEditable)
                self.table.setItem(row, 2, item_path)

                # Status
                item_status = QTableWidgetItem("Pending")
                item_status.setFlags(item_status.flags() & ~Qt.ItemFlag.ItemIsEditable)
                self.table.setItem(row, 3, item_status)

                # Model Role (editable)
                item_role = QTableWidgetItem("")  # auto-filled by worker, user can override
                self.table.setItem(row, 4, item_role)

                # Row Count
                item_rows = QTableWidgetItem("")
                item_rows.setFlags(item_rows.flags() & ~Qt.ItemFlag.ItemIsEditable)
                self.table.setItem(row, 5, item_rows)

                # Column Count
                item_cols = QTableWidgetItem("")
                item_cols.setFlags(item_cols.flags() & ~Qt.ItemFlag.ItemIsEditable)
                self.table.setItem(row, 6, item_cols)

                # Hierarchy Hints (editable if you want to tweak)
                item_hier = QTableWidgetItem("")
                self.table.setItem(row, 7, item_hier)
        except Exception as e:
            self.log(f"ERROR populating table: {e}")

    def on_worker_progress(self, current, total, message):
        try:
            if total > 0:
                perc = int(current * 100 / total)
            else:
                perc = 0
            self.progress_bar.setMaximum(100)
            self.progress_bar.setValue(perc)
            self.status_label.setText(message)
            self.log(f"[{current}/{total}] {message}")
        except Exception as e:
            self.log(f"ERROR updating progress: {e}")

    def on_worker_file_updated(self, index, status, row_count, col_count, model_role, hierarchy_hints):
        try:
            if 0 <= index < self.table.rowCount():
                status_item = self.table.item(index, 3)
                if status_item is not None:
                    status_item.setText(status)

                rows_item = self.table.item(index, 5)
                cols_item = self.table.item(index, 6)
                role_item = self.table.item(index, 4)
                hier_item = self.table.item(index, 7)

                if rows_item is not None and row_count is not None and row_count > 0:
                    rows_item.setText(str(row_count))
                if cols_item is not None and col_count is not None and col_count > 0:
                    cols_item.setText(str(col_count))

                if role_item is None:
                    role_item = QTableWidgetItem("")
                    self.table.setItem(index, 4, role_item)
                if hier_item is None:
                    hier_item = QTableWidgetItem("")
                    self.table.setItem(index, 7, hier_item)

                # Only auto-fill if we actually have a suggestion
                if model_role:
                    role_item.setText(model_role)
                if hierarchy_hints:
                    hier_item.setText(hierarchy_hints)

            self.log(
                f"File #{index + 1} - {status} | rows={row_count}, cols={col_count}, "
                f"auto role='{model_role}', hier='{hierarchy_hints}'"
            )
        except Exception as e:
            self.log(f"ERROR updating file row {index}: {e}")

    def on_worker_finished(self, report_path):
        try:
            self.scan_btn.setEnabled(True)
            self.progress_bar.setValue(100)
            self.status_label.setText("Profiling complete.")
            self.report_path = report_path
            self.open_report_btn.setEnabled(True)
            self.log(f"Profiling complete. Report saved to: {report_path}")
            self.log("Auto Fact/Dim role + Hierarchy hints are now visible in the table and Summary sheet.")

            QMessageBox.information(
                self,
                "Profiling Complete",
                f"Excel report created:\n{report_path}"
            )
        except Exception as e:
            self.log(f"ERROR finishing worker: {e}")

    def on_worker_error(self, msg):
        # Fatal errors only
        self.scan_btn.setEnabled(True)
        self.status_label.setText("Error.")
        self.log(f"ERROR: {msg}")
        QMessageBox.critical(self, "Error", msg)

    def open_report(self):
        if not self.report_path or not os.path.exists(self.report_path):
            QMessageBox.warning(self, "No Report", "Report file not found.")
            return
        try:
            os.startfile(self.report_path)
        except Exception as e:
            self.log(f"ERROR opening report: {e}")
            QMessageBox.critical(self, "Error", f"Cannot open report: {e}")

    def on_cell_double_clicked(self, row, column):
        # If user double-clicks the Model Role or Hierarchy columns, let them edit and do nothing else.
        if column in (4, 7):
            return

        try:
            path_item = self.table.item(row, 2)
            if not path_item:
                return
            path = path_item.text()
            if not os.path.exists(path):
                QMessageBox.warning(self, "File Missing", f"File not found:\n{path}")
                return
            os.startfile(path)
        except Exception as e:
            self.log(f"ERROR opening CSV file: {e}")
            QMessageBox.critical(self, "Error", f"Cannot open file:\n{e}")


# ------------------------------
# Main entry point
# ------------------------------
def main():
    app = QApplication(sys.argv)
    window = CsvProfilerWindow()
    window.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
