# main.py
import csv
import os
import re
import math
from dataclasses import dataclass
from typing import List, Optional, Tuple

from PySide6.QtCore import Qt, QThread, Signal
from PySide6.QtGui import QFont
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QFileDialog, QMessageBox,
    QVBoxLayout, QHBoxLayout, QPushButton, QLabel, QListWidget,
    QCheckBox, QComboBox, QLineEdit, QTabWidget, QGroupBox,
    QFormLayout, QProgressBar, QTextEdit, QSpinBox
)

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# -----------------------------
# Styling (Pastel Themes)
# -----------------------------
@dataclass
class Theme:
    name: str
    header_fill: str
    header_font_color: str
    zebra_fill: str
    tab_colors: List[str]  # ARGB like "FFB7E4C7"


THEMES: List[Theme] = [
    Theme(
        name="Pastel Mint",
        header_fill="FFB7E4C7",
        header_font_color="FF0B3D2E",
        zebra_fill="FFEAF7EF",
        tab_colors=["FFB7E4C7", "FFBDE0FE", "FFFFC8DD", "FFFFE5B4", "FFD8F3DC", "FFCDEAC0"]
    ),
    Theme(
        name="Pastel Lavender",
        header_fill="FFD7C7FF",
        header_font_color="FF2B1B4D",
        zebra_fill="FFF2EEFF",
        tab_colors=["FFD7C7FF", "FFCDB4DB", "FFE0BBE4", "FFFEC8D8", "FFFFDFD3", "FFEAD7FF"]
    ),
    Theme(
        name="Pastel Sky",
        header_fill="FFBDE0FE",
        header_font_color="FF0B2A45",
        zebra_fill="FFEAF3FF",
        tab_colors=["FFBDE0FE", "FFCDF0EA", "FFFFC8DD", "FFFFE5B4", "FFD8F3DC", "FFA2D2FF"]
    ),
    Theme(
        name="Pastel Peach",
        header_fill="FFFFC8B4",
        header_font_color="FF4A1F12",
        zebra_fill="FFFFF1EA",
        tab_colors=["FFFFC8B4", "FFFFE5B4", "FFFFC8DD", "FFBDE0FE", "FFD8F3DC", "FFFFD6A5"]
    ),
]


# -----------------------------
# Parsing / Type inference
# -----------------------------
_NUMERIC_CURRENCY_RE = re.compile(r"^\s*[$€£₹]\s*[-+]?\d[\d,]*([.]\d+)?\s*$")
_NUMERIC_PCT_RE = re.compile(r"^\s*[-+]?\d[\d,]*([.]\d+)?\s*%\s*$")
_NUMERIC_PLAIN_RE = re.compile(r"^\s*[-+]?\d[\d,]*([.]\d+)?\s*$")

DEFAULT_TEXT_COLUMN_HINTS = "id,code,zip,postal,phone,mobile,ssn,pan,gstin,sku,account,acct,invoice,po,order,txn,transaction,ref,reference"


def sanitize_sheet_name(name: str) -> str:
    bad = r'[:\\/?*\[\]]'
    name = re.sub(bad, " ", name).strip()
    if not name:
        name = "Sheet"
    name = name[:31]
    return name


def should_force_text_by_header(header: str, hints: List[str]) -> bool:
    h = (header or "").strip().lower()
    if not h:
        return False
    return any(tok in h for tok in hints)


def looks_like_leading_zero_identifier(s: str) -> bool:
    s = s.strip()
    return len(s) > 1 and s.isdigit() and s[0] == "0"


def parse_cell_value(
    raw: str,
    allow_thousands: bool,
    allow_currency_percent: bool,
    preserve_leading_zeros: bool,
    force_text: bool
) -> Tuple[object, Optional[str]]:
    """
    Returns (value, number_format)
    number_format can be None -> leave as default (General)
    """
    if raw is None:
        return "", None
    s = str(raw)

    if s == "":
        return "", None

    if force_text:
        return s, "@"

    st = s.strip()

    if preserve_leading_zeros and looks_like_leading_zero_identifier(st):
        return st, "@"

    if allow_currency_percent and _NUMERIC_CURRENCY_RE.match(st):
        cleaned = re.sub(r"[$€£₹]", "", st).strip()
        cleaned = cleaned.replace(",", "") if allow_thousands else cleaned
        try:
            val = float(cleaned)
            return val, "#,##0.00"
        except Exception:
            return s, "@"

    if allow_currency_percent and _NUMERIC_PCT_RE.match(st):
        cleaned = st.replace("%", "").strip()
        cleaned = cleaned.replace(",", "") if allow_thousands else cleaned
        try:
            val = float(cleaned) / 100.0
            return val, "0.00%"
        except Exception:
            return s, "@"

    if _NUMERIC_PLAIN_RE.match(st):
        cleaned = st.replace(",", "") if allow_thousands else st
        try:
            if "." in cleaned:
                return float(cleaned), None
            else:
                return int(cleaned), None
        except Exception:
            return s, "@"

    return s, "@"


# -----------------------------
# Excel formatting helpers
# -----------------------------
THIN = Side(style="thin", color="FFB0B0B0")
BORDER_THIN = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def apply_sheet_formatting(
    ws,
    theme: Theme,
    header_row: int = 1,
    wrap_len: int = 40,
    max_col_width: int = 60,
    min_col_width: int = 10,
    zebra: bool = True,
    add_filters: bool = True
):
    header_fill = PatternFill("solid", fgColor=theme.header_fill)
    header_font = Font(bold=True, color=theme.header_font_color)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.freeze_panes = f"A{header_row+1}"

    max_row = ws.max_row
    max_col = ws.max_column

    if max_row >= header_row:
        for c in range(1, max_col + 1):
            cell = ws.cell(row=header_row, column=c)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = BORDER_THIN

    zebra_fill = PatternFill("solid", fgColor=theme.zebra_fill)
    for r in range(header_row + 1, max_row + 1):
        is_zebra = zebra and ((r - (header_row + 1)) % 2 == 0)
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER_THIN
            if isinstance(cell.value, str):
                if len(cell.value) >= wrap_len or ("\n" in cell.value):
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                else:
                    cell.alignment = Alignment(vertical="top", wrap_text=False)
            else:
                cell.alignment = Alignment(vertical="top", wrap_text=False)

            if is_zebra:
                if cell.fill is None or cell.fill.patternType is None:
                    cell.fill = zebra_fill

    if add_filters:
        try:
            if max_row >= header_row and max_col >= 1:
                ref = f"A{header_row}:{get_column_letter(max_col)}{max_row}"
                ws.auto_filter.ref = ref
        except Exception:
            pass

    col_widths = [min_col_width] * (max_col + 1)
    for c in range(1, max_col + 1):
        best = 0
        for r in range(1, max_row + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            if isinstance(v, float):
                s = f"{v:.6g}"
            else:
                s = str(v)
            s = s.replace("\t", " ")
            best = max(best, min(len(s), max_col_width))
        width = max(min_col_width, min(max_col_width, best + 2))
        ws.column_dimensions[get_column_letter(c)].width = width
        col_widths[c] = width

    for r in range(1, max_row + 1):
        max_lines = 1
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            v = cell.value
            if not isinstance(v, str) or not v:
                continue
            if cell.alignment and cell.alignment.wrap_text:
                approx_chars_per_line = max(8, int(col_widths[c] * 1.1))
                chunks = v.splitlines() or [v]
                lines = 0
                for chunk in chunks:
                    lines += max(1, math.ceil(len(chunk) / approx_chars_per_line))
                max_lines = max(max_lines, lines)

        if max_lines > 1:
            ws.row_dimensions[r].height = min(180, 15 * max_lines)


# -----------------------------
# XLSX -> XLSX conversion (format + type fix)
# -----------------------------
def format_xlsx_files(
    xlsx_paths: List[str],
    default_same_dir: bool,
    out_dir: str,
    theme: Theme,
    tab_colorize: bool,
    infer_numbers: bool,
    allow_thousands: bool,
    allow_currency_percent: bool,
    preserve_leading_zeros: bool,
    text_header_hints: List[str],
    wrap_len: int,
    add_filters: bool,
    suffix: str = "_formatted",
    progress_cb=None,
    log_cb=None
):
    def log(msg: str):
        if log_cb:
            log_cb(msg)

    if not default_same_dir:
        os.makedirs(out_dir, exist_ok=True)

    total = len(xlsx_paths)
    for idx, xlsx_path in enumerate(xlsx_paths, start=1):
        log(f"Reading XLSX: {xlsx_path}")
        wb = load_workbook(xlsx_path)

        base = os.path.splitext(os.path.basename(xlsx_path))[0]
        base_dir = os.path.dirname(xlsx_path)

        root_out = base_dir if default_same_dir else out_dir
        os.makedirs(root_out, exist_ok=True)

        out_path = os.path.join(root_out, f"{base}{suffix}.xlsx")

        sheets = wb.sheetnames
        for si, sname in enumerate(sheets, start=1):
            ws = wb[sname]

            # optional tab coloring
            if tab_colorize and theme.tab_colors:
                color = theme.tab_colors[(si - 1) % len(theme.tab_colors)]
                try:
                    ws.sheet_properties.tabColor = color
                except Exception:
                    pass

            # number inference based on header row
            if infer_numbers and ws.max_row >= 1 and ws.max_column >= 1:
                header_vals = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
                header_strs = [(str(v) if v is not None else "") for v in header_vals]

                force_text_cols = set()
                for ci, h in enumerate(header_strs, start=1):
                    if should_force_text_by_header(h, text_header_hints):
                        force_text_cols.add(ci)

                # start from row 2 (assume row 1 is header)
                for r in range(2, ws.max_row + 1):
                    for c in range(1, ws.max_column + 1):
                        cell = ws.cell(row=r, column=c)

                        # keep formulas untouched
                        if isinstance(cell.value, str) and cell.value.lstrip().startswith("="):
                            continue

                        if isinstance(cell.value, str):
                            raw = cell.value
                            force_text = (c in force_text_cols)
                            val, num_fmt = parse_cell_value(
                                raw=raw,
                                allow_thousands=allow_thousands,
                                allow_currency_percent=allow_currency_percent,
                                preserve_leading_zeros=preserve_leading_zeros,
                                force_text=force_text
                            )
                            cell.value = val
                            if num_fmt:
                                cell.number_format = num_fmt

            # apply formatting
            apply_sheet_formatting(ws, theme, wrap_len=wrap_len, add_filters=add_filters)

        wb.save(out_path)
        log(f"Saved: {out_path}")

        if progress_cb:
            progress_cb(int(idx * 100 / total))


# -----------------------------
# Worker thread
# -----------------------------
class Worker(QThread):
    progress = Signal(int)
    log = Signal(str)
    done = Signal(bool, str)

    def __init__(self, fn, *args, **kwargs):
        super().__init__()
        self.fn = fn
        self.args = args
        self.kwargs = kwargs

    def run(self):
        try:
            self.fn(*self.args, **self.kwargs)
            self.done.emit(True, "Completed successfully.")
        except Exception as e:
            self.done.emit(False, f"Failed: {e}")


# -----------------------------
# Conversion logic (existing)
# -----------------------------
def csv_to_xlsx(
    csv_paths: List[str],
    out_dir: str,
    make_single_workbook: bool,
    theme: Theme,
    tab_colorize: bool,
    delimiter: str,
    encoding: str,
    infer_numbers: bool,
    allow_thousands: bool,
    allow_currency_percent: bool,
    preserve_leading_zeros: bool,
    text_header_hints: List[str],
    wrap_len: int,
    add_filters: bool,
    progress_cb=None,
    log_cb=None
):
    def log(msg: str):
        if log_cb:
            log_cb(msg)

    os.makedirs(out_dir, exist_ok=True)

    if make_single_workbook:
        wb = Workbook()
        default_ws = wb.active
        wb.remove(default_ws)

    total = len(csv_paths)
    for idx, csv_path in enumerate(csv_paths, start=1):
        base = os.path.splitext(os.path.basename(csv_path))[0]
        sheet_name = sanitize_sheet_name(base)

        if make_single_workbook:
            ws = wb.create_sheet(title=sheet_name)
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name

        if tab_colorize and theme.tab_colors:
            color = theme.tab_colors[(idx - 1) % len(theme.tab_colors)]
            try:
                ws.sheet_properties.tabColor = color
            except Exception:
                pass

        log(f"Reading CSV: {csv_path}")

        with open(csv_path, "r", newline="", encoding=encoding, errors="replace") as f:
            reader = csv.reader(f, delimiter=delimiter)
            rows = list(reader)

        if not rows:
            log(f"  (empty) {csv_path}")
            apply_sheet_formatting(ws, theme, wrap_len=wrap_len, add_filters=add_filters)
        else:
            header = rows[0]
            ws.append(header)

            force_text_cols = set()
            for ci, h in enumerate(header, start=1):
                if should_force_text_by_header(h, text_header_hints):
                    force_text_cols.add(ci)

            for r in rows[1:]:
                if len(r) < len(header):
                    r = r + [""] * (len(header) - len(r))

                ws.append(r)

                if infer_numbers:
                    row_idx = ws.max_row
                    for ci in range(1, len(r) + 1):
                        raw = r[ci - 1]
                        force_text = ci in force_text_cols
                        val, num_fmt = parse_cell_value(
                            raw=raw,
                            allow_thousands=allow_thousands,
                            allow_currency_percent=allow_currency_percent,
                            preserve_leading_zeros=preserve_leading_zeros,
                            force_text=force_text
                        )
                        cell = ws.cell(row=row_idx, column=ci)
                        cell.value = val
                        if num_fmt:
                            cell.number_format = num_fmt

        apply_sheet_formatting(ws, theme, wrap_len=wrap_len, add_filters=add_filters)

        if not make_single_workbook:
            out_path = os.path.join(out_dir, f"{base}.xlsx")
            wb.save(out_path)
            log(f"Saved: {out_path}")

        if progress_cb:
            progress_cb(int(idx * 100 / total))

    if make_single_workbook:
        out_path = os.path.join(out_dir, "combined.xlsx")
        wb.save(out_path)
        log(f"Saved combined workbook: {out_path}")


def xlsx_to_csv(
    xlsx_paths: List[str],
    default_same_dir: bool,
    out_dir: str,
    per_workbook_folder: bool,
    delimiter: str,
    encoding: str,
    progress_cb=None,
    log_cb=None
):
    def log(msg: str):
        if log_cb:
            log_cb(msg)

    total = len(xlsx_paths)
    for idx, xlsx_path in enumerate(xlsx_paths, start=1):
        log(f"Reading XLSX: {xlsx_path}")
        wb = load_workbook(xlsx_path, data_only=True)

        xlsx_base = os.path.splitext(os.path.basename(xlsx_path))[0]
        base_dir = os.path.dirname(xlsx_path)

        if default_same_dir:
            root_out = base_dir
        else:
            root_out = out_dir
            os.makedirs(root_out, exist_ok=True)

        if per_workbook_folder:
            root_out = os.path.join(root_out, xlsx_base)
            os.makedirs(root_out, exist_ok=True)

        sheets = wb.sheetnames
        is_single_sheet = len(sheets) == 1

        for sname in sheets:
            ws = wb[sname]
            safe_sheet = sanitize_sheet_name(sname)

            if is_single_sheet:
                csv_name = f"{xlsx_base}.csv"
            else:
                csv_name = f"{safe_sheet}.csv"

            csv_path = os.path.join(root_out, csv_name)

            with open(csv_path, "w", newline="", encoding=encoding, errors="replace") as f:
                writer = csv.writer(f, delimiter=delimiter)
                for row in ws.iter_rows(values_only=True):
                    out_row = []
                    for v in row:
                        if v is None:
                            out_row.append("")
                        else:
                            out_row.append(str(v))
                    writer.writerow(out_row)

            log(f"Saved: {csv_path}")

        if progress_cb:
            progress_cb(int(idx * 100 / total))


# -----------------------------
# Qt UI
# -----------------------------
PASTEL_APP_STYLESHEET = """
QPushButton {
    background-color: #BDE0FE;
    color: #102A43;
    border: 1px solid #A7C8E8;
    border-radius: 8px;
    padding: 7px 12px;
    font-weight: 600;
}
QPushButton:hover {
    background-color: #CDF0EA;
}
QPushButton:pressed {
    background-color: #B7E4C7;
}
QPushButton:disabled {
    background-color: #E6EEF7;
    color: #7A8699;
    border: 1px solid #D6DEE8;
}
QGroupBox {
    font-weight: 700;
}
QLineEdit, QComboBox, QSpinBox, QListWidget, QTextEdit {
    border: 1px solid #D6DEE8;
    border-radius: 6px;
    padding: 4px;
}
QTabWidget::pane {
    border: 1px solid #D6DEE8;
    border-radius: 8px;
}
"""


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("CSV ↔ XLSX Processor (Qt6)")
        self.setMinimumSize(1100, 760)
        self.setStyleSheet(PASTEL_APP_STYLESHEET)

        root = QWidget()
        self.setCentralWidget(root)
        main_layout = QVBoxLayout(root)

        title = QLabel("CSV ↔ XLSX Processor")
        title.setFont(QFont("Segoe UI", 16, QFont.Weight.Bold))
        main_layout.addWidget(title)

        self.tabs = QTabWidget()
        main_layout.addWidget(self.tabs, 2)

        self.log_box = QTextEdit()
        self.log_box.setReadOnly(True)
        self.log_box.setPlaceholderText("Run logs will appear here...")
        main_layout.addWidget(self.log_box, 1)

        bottom = QHBoxLayout()
        self.progress = QProgressBar()
        self.progress.setValue(0)
        bottom.addWidget(self.progress, 1)

        self.run_btn = QPushButton("Run")
        self.run_btn.clicked.connect(self.on_run)
        bottom.addWidget(self.run_btn)

        main_layout.addLayout(bottom)

        self.worker: Optional[Worker] = None

        self._build_tab_csv_to_xlsx()
        self._build_tab_xlsx_to_csv()
        self._build_tab_xlsx_to_xlsx_format()

    def _build_tab_csv_to_xlsx(self):
        w = QWidget()
        layout = QVBoxLayout(w)

        pick_row = QHBoxLayout()
        self.csv_list = QListWidget()
        pick_row.addWidget(self.csv_list, 1)

        btn_col = QVBoxLayout()
        add_btn = QPushButton("Add CSV files…")
        add_btn.clicked.connect(self.add_csv_files)
        btn_col.addWidget(add_btn)

        rm_btn = QPushButton("Remove selected")
        rm_btn.clicked.connect(lambda: self._remove_selected(self.csv_list))
        btn_col.addWidget(rm_btn)

        clr_btn = QPushButton("Clear")
        clr_btn.clicked.connect(self.csv_list.clear)
        btn_col.addWidget(clr_btn)

        btn_col.addStretch(1)
        pick_row.addLayout(btn_col)
        layout.addLayout(pick_row, 1)

        grp = QGroupBox("CSV → XLSX Options (Excel correctness + usability)")
        form = QFormLayout(grp)

        self.csv_out_dir = QLineEdit()
        self.csv_out_dir.setPlaceholderText("Output folder for generated XLSX files")
        out_btn = QPushButton("Browse…")
        out_btn.clicked.connect(self.pick_csv_out_dir)
        out_row = QHBoxLayout()
        out_row.addWidget(self.csv_out_dir, 1)
        out_row.addWidget(out_btn)
        out_wrap = QWidget()
        out_wrap.setLayout(out_row)
        form.addRow("Output folder:", out_wrap)

        self.make_single_wb = QCheckBox("Combine into ONE workbook (each CSV becomes a sheet)")
        self.make_single_wb.setChecked(False)
        form.addRow("", self.make_single_wb)

        self.tab_colorize = QCheckBox("Color worksheet tabs (pastel)")
        self.tab_colorize.setChecked(True)
        form.addRow("", self.tab_colorize)

        self.theme_combo = QComboBox()
        for t in THEMES:
            self.theme_combo.addItem(t.name)
        self.theme_combo.setCurrentIndex(0)
        form.addRow("Workbook style:", self.theme_combo)

        self.delim_csv = QLineEdit(",")
        self.delim_csv.setMaxLength(3)
        form.addRow("CSV delimiter:", self.delim_csv)

        self.encoding_csv = QLineEdit("utf-8")
        form.addRow("CSV encoding:", self.encoding_csv)

        self.infer_numbers = QCheckBox("Infer numbers to avoid Excel 'Number stored as text' warnings (Recommended)")
        self.infer_numbers.setChecked(True)
        form.addRow("", self.infer_numbers)

        self.allow_thousands = QCheckBox("Accept thousands separators (1,234)")
        self.allow_thousands.setChecked(True)
        form.addRow("", self.allow_thousands)

        self.allow_currency_percent = QCheckBox("Accept currency/percent patterns ($1,234.50 / 12.5%)")
        self.allow_currency_percent.setChecked(True)
        form.addRow("", self.allow_currency_percent)

        self.preserve_leading_zeros = QCheckBox("Preserve leading-zero identifiers (e.g., 00123) as text")
        self.preserve_leading_zeros.setChecked(True)
        form.addRow("", self.preserve_leading_zeros)

        self.text_hints = QLineEdit(DEFAULT_TEXT_COLUMN_HINTS)
        form.addRow("Force-text header hints:", self.text_hints)

        self.add_filters = QCheckBox("Apply AutoFilter on header row (best-effort)")
        self.add_filters.setChecked(True)
        form.addRow("", self.add_filters)

        self.wrap_len = QSpinBox()
        self.wrap_len.setRange(10, 200)
        self.wrap_len.setValue(40)
        form.addRow("Wrap text when length ≥", self.wrap_len)

        layout.addWidget(grp)
        self.tabs.addTab(w, "CSV → XLSX")

    def _build_tab_xlsx_to_csv(self):
        w = QWidget()
        layout = QVBoxLayout(w)

        pick_row = QHBoxLayout()
        self.xlsx_list = QListWidget()
        pick_row.addWidget(self.xlsx_list, 1)

        btn_col = QVBoxLayout()
        add_btn = QPushButton("Add XLSX files…")
        add_btn.clicked.connect(self.add_xlsx_files)
        btn_col.addWidget(add_btn)

        rm_btn = QPushButton("Remove selected")
        rm_btn.clicked.connect(lambda: self._remove_selected(self.xlsx_list))
        btn_col.addWidget(rm_btn)

        clr_btn = QPushButton("Clear")
        clr_btn.clicked.connect(self.xlsx_list.clear)
        btn_col.addWidget(clr_btn)

        btn_col.addStretch(1)
        pick_row.addLayout(btn_col)
        layout.addLayout(pick_row, 1)

        grp = QGroupBox("XLSX → CSV Options")
        form = QFormLayout(grp)

        self.xlsx_default_same_dir = QCheckBox("Write CSVs into the same folder as the source XLSX (Default)")
        self.xlsx_default_same_dir.setChecked(True)
        self.xlsx_default_same_dir.stateChanged.connect(self._toggle_xlsx_outdir_enabled)
        form.addRow("", self.xlsx_default_same_dir)

        self.xlsx_out_dir = QLineEdit()
        self.xlsx_out_dir.setPlaceholderText("Output folder (used only if 'same folder' is unchecked)")
        out_btn = QPushButton("Browse…")
        out_btn.clicked.connect(self.pick_xlsx_out_dir)

        out_row = QHBoxLayout()
        out_row.addWidget(self.xlsx_out_dir, 1)
        out_row.addWidget(out_btn)
        out_wrap = QWidget()
        out_wrap.setLayout(out_row)
        form.addRow("Output folder:", out_wrap)

        self.per_workbook_folder = QCheckBox("Create a separate folder per XLSX workbook (folder name = XLSX name)")
        self.per_workbook_folder.setChecked(False)
        form.addRow("", self.per_workbook_folder)

        self.delim_xlsx = QLineEdit(",")
        self.delim_xlsx.setMaxLength(3)
        form.addRow("CSV delimiter:", self.delim_xlsx)

        self.encoding_xlsx = QLineEdit("utf-8")
        form.addRow("CSV encoding:", self.encoding_xlsx)

        layout.addWidget(grp)
        self._toggle_xlsx_outdir_enabled()
        self.tabs.addTab(w, "XLSX → CSV")

    def _build_tab_xlsx_to_xlsx_format(self):
        w = QWidget()
        layout = QVBoxLayout(w)

        pick_row = QHBoxLayout()
        self.xlsx_fmt_list = QListWidget()
        pick_row.addWidget(self.xlsx_fmt_list, 1)

        btn_col = QVBoxLayout()
        add_btn = QPushButton("Add XLSX files…")
        add_btn.clicked.connect(self.add_xlsx_fmt_files)
        btn_col.addWidget(add_btn)

        rm_btn = QPushButton("Remove selected")
        rm_btn.clicked.connect(lambda: self._remove_selected(self.xlsx_fmt_list))
        btn_col.addWidget(rm_btn)

        clr_btn = QPushButton("Clear")
        clr_btn.clicked.connect(self.xlsx_fmt_list.clear)
        btn_col.addWidget(clr_btn)

        btn_col.addStretch(1)
        pick_row.addLayout(btn_col)
        layout.addLayout(pick_row, 1)

        grp = QGroupBox("XLSX → XLSX Formatting Options (make Excel look clean + fix numeric text)")
        form = QFormLayout(grp)

        self.xlsx_fmt_same_dir = QCheckBox("Save formatted XLSX next to source file (Default)")
        self.xlsx_fmt_same_dir.setChecked(True)
        self.xlsx_fmt_same_dir.stateChanged.connect(self._toggle_xlsx_fmt_outdir_enabled)
        form.addRow("", self.xlsx_fmt_same_dir)

        self.xlsx_fmt_out_dir = QLineEdit()
        self.xlsx_fmt_out_dir.setPlaceholderText("Output folder (used only if 'same folder' is unchecked)")
        out_btn = QPushButton("Browse…")
        out_btn.clicked.connect(self.pick_xlsx_fmt_out_dir)

        out_row = QHBoxLayout()
        out_row.addWidget(self.xlsx_fmt_out_dir, 1)
        out_row.addWidget(out_btn)
        out_wrap = QWidget()
        out_wrap.setLayout(out_row)
        form.addRow("Output folder:", out_wrap)

        self.xlsx_fmt_theme = QComboBox()
        for t in THEMES:
            self.xlsx_fmt_theme.addItem(t.name)
        self.xlsx_fmt_theme.setCurrentIndex(0)
        form.addRow("Workbook style:", self.xlsx_fmt_theme)

        self.xlsx_fmt_tab_colorize = QCheckBox("Color worksheet tabs (pastel)")
        self.xlsx_fmt_tab_colorize.setChecked(True)
        form.addRow("", self.xlsx_fmt_tab_colorize)

        self.xlsx_fmt_infer_numbers = QCheckBox("Convert numeric-looking text into real numbers (Recommended)")
        self.xlsx_fmt_infer_numbers.setChecked(True)
        form.addRow("", self.xlsx_fmt_infer_numbers)

        self.xlsx_fmt_allow_thousands = QCheckBox("Accept thousands separators (1,234)")
        self.xlsx_fmt_allow_thousands.setChecked(True)
        form.addRow("", self.xlsx_fmt_allow_thousands)

        self.xlsx_fmt_allow_currency_percent = QCheckBox("Accept currency/percent patterns ($1,234.50 / 12.5%)")
        self.xlsx_fmt_allow_currency_percent.setChecked(True)
        form.addRow("", self.xlsx_fmt_allow_currency_percent)

        self.xlsx_fmt_preserve_leading_zeros = QCheckBox("Preserve leading-zero identifiers (e.g., 00123) as text")
        self.xlsx_fmt_preserve_leading_zeros.setChecked(True)
        form.addRow("", self.xlsx_fmt_preserve_leading_zeros)

        self.xlsx_fmt_text_hints = QLineEdit(DEFAULT_TEXT_COLUMN_HINTS)
        form.addRow("Force-text header hints:", self.xlsx_fmt_text_hints)

        self.xlsx_fmt_add_filters = QCheckBox("Apply AutoFilter on header row (best-effort)")
        self.xlsx_fmt_add_filters.setChecked(True)
        form.addRow("", self.xlsx_fmt_add_filters)

        self.xlsx_fmt_wrap_len = QSpinBox()
        self.xlsx_fmt_wrap_len.setRange(10, 200)
        self.xlsx_fmt_wrap_len.setValue(40)
        form.addRow("Wrap text when length ≥", self.xlsx_fmt_wrap_len)

        layout.addWidget(grp)
        self._toggle_xlsx_fmt_outdir_enabled()

        self.tabs.addTab(w, "XLSX → XLSX (Format)")

    def _toggle_xlsx_outdir_enabled(self):
        enabled = not self.xlsx_default_same_dir.isChecked()
        self.xlsx_out_dir.setEnabled(enabled)

    def _toggle_xlsx_fmt_outdir_enabled(self):
        enabled = not self.xlsx_fmt_same_dir.isChecked()
        self.xlsx_fmt_out_dir.setEnabled(enabled)

    def _remove_selected(self, lw: QListWidget):
        for item in lw.selectedItems():
            lw.takeItem(lw.row(item))

    def add_csv_files(self):
        paths, _ = QFileDialog.getOpenFileNames(self, "Select CSV files", "", "CSV Files (*.csv);;All Files (*.*)")
        for p in paths:
            self.csv_list.addItem(p)

    def add_xlsx_files(self):
        paths, _ = QFileDialog.getOpenFileNames(self, "Select XLSX files", "", "Excel Files (*.xlsx);;All Files (*.*)")
        for p in paths:
            self.xlsx_list.addItem(p)

    def add_xlsx_fmt_files(self):
        paths, _ = QFileDialog.getOpenFileNames(self, "Select XLSX files to format", "", "Excel Files (*.xlsx);;All Files (*.*)")
        for p in paths:
            self.xlsx_fmt_list.addItem(p)

    def pick_csv_out_dir(self):
        d = QFileDialog.getExistingDirectory(self, "Choose output folder")
        if d:
            self.csv_out_dir.setText(d)

    def pick_xlsx_out_dir(self):
        d = QFileDialog.getExistingDirectory(self, "Choose output folder")
        if d:
            self.xlsx_out_dir.setText(d)

    def pick_xlsx_fmt_out_dir(self):
        d = QFileDialog.getExistingDirectory(self, "Choose output folder")
        if d:
            self.xlsx_fmt_out_dir.setText(d)

    def log(self, msg: str):
        self.log_box.append(msg)

    def set_progress(self, v: int):
        self.progress.setValue(max(0, min(100, v)))

    def on_run(self):
        self.log_box.clear()
        self.set_progress(0)

        tab_idx = self.tabs.currentIndex()
        if tab_idx == 0:
            self.run_csv_to_xlsx()
        elif tab_idx == 1:
            self.run_xlsx_to_csv()
        else:
            self.run_xlsx_to_xlsx_format()

    def run_csv_to_xlsx(self):
        csv_paths = [self.csv_list.item(i).text() for i in range(self.csv_list.count())]
        if not csv_paths:
            QMessageBox.warning(self, "Missing input", "Please add at least one CSV file.")
            return

        out_dir = self.csv_out_dir.text().strip()
        if not out_dir:
            QMessageBox.warning(self, "Missing output", "Please select an output folder for XLSX generation.")
            return

        theme = THEMES[self.theme_combo.currentIndex()]
        delimiter = self.delim_csv.text()
        if delimiter == r"\t":
            delimiter = "\t"
        if not delimiter:
            delimiter = ","

        encoding = self.encoding_csv.text().strip() or "utf-8"
        hints = [h.strip().lower() for h in self.text_hints.text().split(",") if h.strip()]

        self.log("CSV → XLSX job started")
        self.log(f"- Files: {len(csv_paths)}")
        self.log(f"- Output: {out_dir}")
        self.log(f"- Mode: {'Single workbook' if self.make_single_wb.isChecked() else 'One XLSX per CSV'}")
        self.log(f"- Style: {theme.name}")

        def job():
            csv_to_xlsx(
                csv_paths=csv_paths,
                out_dir=out_dir,
                make_single_workbook=self.make_single_wb.isChecked(),
                theme=theme,
                tab_colorize=self.tab_colorize.isChecked(),
                delimiter=delimiter,
                encoding=encoding,
                infer_numbers=self.infer_numbers.isChecked(),
                allow_thousands=self.allow_thousands.isChecked(),
                allow_currency_percent=self.allow_currency_percent.isChecked(),
                preserve_leading_zeros=self.preserve_leading_zeros.isChecked(),
                text_header_hints=hints,
                wrap_len=self.wrap_len.value(),
                add_filters=self.add_filters.isChecked(),
                progress_cb=self.set_progress,
                log_cb=self.log
            )

        self._start_worker(job)

    def run_xlsx_to_csv(self):
        xlsx_paths = [self.xlsx_list.item(i).text() for i in range(self.xlsx_list.count())]
        if not xlsx_paths:
            QMessageBox.warning(self, "Missing input", "Please add at least one XLSX file.")
            return

        same_dir = self.xlsx_default_same_dir.isChecked()
        out_dir = self.xlsx_out_dir.text().strip()
        if not same_dir and not out_dir:
            QMessageBox.warning(self, "Missing output", "Please select an output folder or enable 'same folder'.")
            return

        delimiter = self.delim_xlsx.text()
        if delimiter == r"\t":
            delimiter = "\t"
        if not delimiter:
            delimiter = ","

        encoding = self.encoding_xlsx.text().strip() or "utf-8"

        self.log("XLSX → CSV job started")
        self.log(f"- Files: {len(xlsx_paths)}")
        self.log(f"- Output: {'Same folder as XLSX' if same_dir else out_dir}")
        self.log(f"- Per-workbook folder: {self.per_workbook_folder.isChecked()}")

        def job():
            xlsx_to_csv(
                xlsx_paths=xlsx_paths,
                default_same_dir=same_dir,
                out_dir=out_dir,
                per_workbook_folder=self.per_workbook_folder.isChecked(),
                delimiter=delimiter,
                encoding=encoding,
                progress_cb=self.set_progress,
                log_cb=self.log
            )

        self._start_worker(job)

    def run_xlsx_to_xlsx_format(self):
        xlsx_paths = [self.xlsx_fmt_list.item(i).text() for i in range(self.xlsx_fmt_list.count())]
        if not xlsx_paths:
            QMessageBox.warning(self, "Missing input", "Please add at least one XLSX file to format.")
            return

        same_dir = self.xlsx_fmt_same_dir.isChecked()
        out_dir = self.xlsx_fmt_out_dir.text().strip()
        if not same_dir and not out_dir:
            QMessageBox.warning(self, "Missing output", "Please select an output folder or enable 'same folder'.")
            return

        theme = THEMES[self.xlsx_fmt_theme.currentIndex()]
        hints = [h.strip().lower() for h in self.xlsx_fmt_text_hints.text().split(",") if h.strip()]

        self.log("XLSX → XLSX (Format) job started")
        self.log(f"- Files: {len(xlsx_paths)}")
        self.log(f"- Output: {'Same folder as XLSX' if same_dir else out_dir}")
        self.log(f"- Style: {theme.name}")
        self.log(f"- Convert numeric text: {self.xlsx_fmt_infer_numbers.isChecked()} | Preserve leading zeros: {self.xlsx_fmt_preserve_leading_zeros.isChecked()}")

        def job():
            format_xlsx_files(
                xlsx_paths=xlsx_paths,
                default_same_dir=same_dir,
                out_dir=out_dir,
                theme=theme,
                tab_colorize=self.xlsx_fmt_tab_colorize.isChecked(),
                infer_numbers=self.xlsx_fmt_infer_numbers.isChecked(),
                allow_thousands=self.xlsx_fmt_allow_thousands.isChecked(),
                allow_currency_percent=self.xlsx_fmt_allow_currency_percent.isChecked(),
                preserve_leading_zeros=self.xlsx_fmt_preserve_leading_zeros.isChecked(),
                text_header_hints=hints,
                wrap_len=self.xlsx_fmt_wrap_len.value(),
                add_filters=self.xlsx_fmt_add_filters.isChecked(),
                suffix="_formatted",
                progress_cb=self.set_progress,
                log_cb=self.log
            )

        self._start_worker(job)

    def _start_worker(self, job_fn):
        self.run_btn.setEnabled(False)

        self.worker = Worker(job_fn)
        self.worker.progress.connect(self.set_progress)
        self.worker.log.connect(self.log)
        self.worker.done.connect(self._on_done)
        self.worker.start()

    def _on_done(self, ok: bool, msg: str):
        self.run_btn.setEnabled(True)
        self.set_progress(100 if ok else self.progress.value())
        if ok:
            QMessageBox.information(self, "Done", msg)
        else:
            QMessageBox.critical(self, "Error", msg)


def main():
    app = QApplication([])
    win = MainWindow()
    win.show()
    app.exec()


if __name__ == "__main__":
    main()
