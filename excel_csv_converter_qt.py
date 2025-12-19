import os
import re
import sys
import time
import traceback
from dataclasses import dataclass
from typing import List, Optional, Tuple

import pandas as pd
from PyQt6.QtCore import Qt, QThread, pyqtSignal, QSettings
from PyQt6.QtGui import QDragEnterEvent, QDropEvent
from PyQt6.QtWidgets import (
    QApplication, QComboBox, QFileDialog, QGroupBox, QHBoxLayout, QLabel,
    QLineEdit, QListWidget, QListWidgetItem, QMainWindow, QMessageBox,
    QPushButton, QProgressBar, QSpinBox, QTabWidget, QVBoxLayout,
    QWidget, QCheckBox, QPlainTextEdit
)

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


INVALID_SHEET_CHARS = r'[:\\/?*\[\]]'
CONFLICT_POLICIES = ["Skip existing", "Overwrite", "Rename (add _1, _2 ...)"]


def ensure_dir(path: str) -> None:
    os.makedirs(path, exist_ok=True)


def base_name_no_ext(path: str) -> str:
    return os.path.splitext(os.path.basename(path))[0]


def is_xlsx(path: str) -> bool:
    return path.lower().endswith(".xlsx")


def is_csv(path: str) -> bool:
    return path.lower().endswith(".csv")


def safe_sheet_name(name: str, fallback: str = "Sheet") -> str:
    name = re.sub(INVALID_SHEET_CHARS, "_", name).strip()
    if not name:
        name = fallback
    return name[:31]


def unique_path(path: str) -> str:
    if not os.path.exists(path):
        return path
    root, ext = os.path.splitext(path)
    i = 1
    while True:
        candidate = f"{root}_{i}{ext}"
        if not os.path.exists(candidate):
            return candidate
        i += 1


def guess_delimiter(delim_choice: str) -> str:
    mapping = {
        "Comma (,)": ",",
        "Tab (\\t)": "\t",
        "Semicolon (;)": ";",
        "Pipe (|)": "|",
    }
    return mapping.get(delim_choice, ",")


def read_csv_safely(csv_path: str, delimiter: str, encoding: str, dtype_mode: str) -> pd.DataFrame:
    if dtype_mode == "All as text":
        return pd.read_csv(csv_path, sep=delimiter, encoding=encoding, dtype=str, keep_default_na=False)
    return pd.read_csv(csv_path, sep=delimiter, encoding=encoding)


def apply_excel_formatting(ws, df: pd.DataFrame, max_col_width: int, wrap_text: bool) -> None:
    header_fill = PatternFill("solid", fgColor="EDEDED")
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap_text)
    cell_alignment = Alignment(vertical="top", wrap_text=wrap_text)

    ws.freeze_panes = "A2"

    for col_idx, col_name in enumerate(df.columns, start=1):
        c = ws.cell(row=1, column=col_idx)
        c.fill = header_fill
        c.font = header_font
        c.alignment = header_alignment

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = cell_alignment

    ws.auto_filter.ref = ws.dimensions

    for col_idx, col_name in enumerate(df.columns, start=1):
        letter = get_column_letter(col_idx)
        max_len = len(str(col_name)) if col_name is not None else 0
        sample_n = min(500, len(df))
        if sample_n > 0:
            series = df.iloc[:sample_n, col_idx - 1].astype(str)
            max_len = max(max_len, int(series.map(len).max()))
        width = min(max_len + 2, max_col_width)
        width = max(10, width)
        ws.column_dimensions[letter].width = width


def write_df_to_sheet(ws, df: pd.DataFrame) -> None:
    ws.append(list(df.columns))
    for row in df.itertuples(index=False, name=None):
        ws.append(list(row))


def handle_conflict(path: str, policy: str) -> Tuple[str, bool]:
    if not os.path.exists(path):
        return path, True
    if policy == "Skip existing":
        return path, False
    if policy == "Overwrite":
        return path, True
    return unique_path(path), True


@dataclass
class XlsxToCsvOptions:
    output_root: str
    output_mode: str
    delimiter: str
    encoding: str
    include_index: bool
    first_row_header: bool
    include_hidden_sheets: bool
    conflict_policy: str


@dataclass
class CsvToXlsxOptions:
    output_root: str
    output_mode: str
    merged_workbook_name: str
    delimiter: str
    encoding: str
    dtype_mode: str
    max_col_width: int
    wrap_text: bool
    conflict_policy: str


@dataclass
class RunStats:
    inputs: int = 0
    outputs_created: int = 0
    skipped: int = 0
    errors: int = 0
    start_ts: float = 0.0
    end_ts: float = 0.0

    def duration_sec(self) -> float:
        return max(0.0, self.end_ts - self.start_ts)


class ConvertWorker(QThread):
    progress = pyqtSignal(int, int)
    status = pyqtSignal(str)
    log = pyqtSignal(str)
    error_item = pyqtSignal(str)
    done = pyqtSignal(object)

    def __init__(self, mode: str, paths: List[str], x2c: Optional[XlsxToCsvOptions], c2x: Optional[CsvToXlsxOptions]):
        super().__init__()
        self.mode = mode
        self.paths = paths
        self.x2c = x2c
        self.c2x = c2x
        self._cancel = False

    def cancel(self):
        self._cancel = True

    def run(self):
        stats = RunStats(inputs=len(self.paths), start_ts=time.time())
        try:
            root = self.x2c.output_root if self.mode == "xlsx_to_csv" else self.c2x.output_root
            ensure_dir(root)  # (1) create output folder if missing

            if self.mode == "xlsx_to_csv":
                self._run_xlsx_to_csv(stats)
            else:
                self._run_csv_to_xlsx(stats)

        except Exception as e:
            stats.errors += 1
            self.log.emit(f"FATAL: {e}")
            self.log.emit(traceback.format_exc())
        finally:
            stats.end_ts = time.time()
            self.done.emit(stats)

    def _is_per_workbook_mode(self, text: str) -> bool:
        # robust check (fixes any minor label mismatch)
        t = (text or "").strip().lower()
        return "per" in t and "workbook" in t

    def _run_xlsx_to_csv(self, stats: RunStats):
        opts = self.x2c
        total = len(self.paths)
        ensure_dir(opts.output_root)

        for i, xlsx_path in enumerate(self.paths, start=1):
            if self._cancel:
                self.log.emit("Cancelled by user.")
                break

            self.progress.emit(i, total)
            self.status.emit(f"Reading: {xlsx_path}")

            wb_name = base_name_no_ext(xlsx_path)

            try:
                wb = load_workbook(xlsx_path, data_only=True)

                sheet_names = wb.sheetnames
                if not opts.include_hidden_sheets:
                    visible = []
                    for sname in sheet_names:
                        sh = wb[sname]
                        if getattr(sh, "sheet_state", "visible") == "visible":
                            visible.append(sname)
                    sheet_names = visible

                if not sheet_names:
                    stats.skipped += 1
                    self.log.emit(f"SKIP (no visible sheets): {xlsx_path}")
                    continue

                # (FIX) Decide folder, then ALWAYS ensure it exists
                if self._is_per_workbook_mode(opts.output_mode):
                    target_folder = os.path.join(opts.output_root, wb_name)
                else:
                    target_folder = opts.output_root

                # (FIX) Always create the target folder here
                ensure_dir(target_folder)
                self.log.emit(f"Target folder: {target_folder}")

                if len(sheet_names) == 1:
                    sname = sheet_names[0]
                    out_csv = os.path.join(target_folder, f"{wb_name}.csv")
                    out_csv, should_write = handle_conflict(out_csv, opts.conflict_policy)
                    if not should_write:
                        stats.skipped += 1
                        self.log.emit(f"SKIP (exists): {out_csv}")
                        continue

                    self.status.emit(f"Exporting sheet '{sname}' → {os.path.basename(out_csv)}")
                    self._export_sheet_to_csv(wb, sname, out_csv, opts)
                    stats.outputs_created += 1
                    self.log.emit(f"OK: {out_csv}")
                else:
                    for sname in sheet_names:
                        if self._cancel:
                            self.log.emit("Cancelled by user.")
                            break

                        safe = safe_sheet_name(sname, "Sheet")
                        out_csv = os.path.join(target_folder, f"{safe}.csv")
                        out_csv, should_write = handle_conflict(out_csv, opts.conflict_policy)
                        if not should_write:
                            stats.skipped += 1
                            self.log.emit(f"SKIP (exists): {out_csv}")
                            continue

                        self.status.emit(f"Exporting sheet '{sname}' → {os.path.basename(out_csv)}")
                        self._export_sheet_to_csv(wb, sname, out_csv, opts)
                        stats.outputs_created += 1
                        self.log.emit(f"OK: {out_csv}")

            except Exception as e:
                stats.errors += 1
                msg = f"ERROR: {xlsx_path} | {e}"
                self.log.emit(msg)
                self.error_item.emit(msg)

    def _export_sheet_to_csv(self, wb, sheet_name: str, out_csv: str, opts: XlsxToCsvOptions) -> None:
        sheet = wb[sheet_name]
        df = pd.DataFrame(sheet.values)

        if df.empty:
            pd.DataFrame().to_csv(out_csv, index=opts.include_index, sep=opts.delimiter, encoding=opts.encoding)
            return

        if opts.first_row_header and len(df) >= 1:
            df.columns = df.iloc[0].astype(str).tolist()
            df = df.iloc[1:].reset_index(drop=True)
        else:
            df.columns = [f"col_{i}" for i in range(1, df.shape[1] + 1)]

        df.to_csv(out_csv, index=opts.include_index, sep=opts.delimiter, encoding=opts.encoding)

    def _run_csv_to_xlsx(self, stats: RunStats):
        opts = self.c2x
        total = len(self.paths)
        ensure_dir(opts.output_root)

        if opts.output_mode == "Merge into one workbook":
            out_path = os.path.join(opts.output_root, f"{base_name_no_ext(opts.merged_workbook_name)}.xlsx")
            out_path, should_write = handle_conflict(out_path, opts.conflict_policy)
            if not should_write:
                stats.skipped += 1
                self.log.emit(f"SKIP (exists): {out_path}")
                return

            self.status.emit("Creating merged workbook…")
            wb = Workbook()
            default = wb.active
            wb.remove(default)

            used_names = set()
            for i, csv_path in enumerate(self.paths, start=1):
                if self._cancel:
                    self.log.emit("Cancelled by user.")
                    break

                self.progress.emit(i, total)
                self.status.emit(f"Reading CSV: {csv_path}")

                try:
                    df = read_csv_safely(csv_path, opts.delimiter, opts.encoding, opts.dtype_mode)
                    name = safe_sheet_name(base_name_no_ext(csv_path), "Sheet")

                    base = name
                    k = 1
                    while name in used_names:
                        suffix = f"_{k}"
                        name = safe_sheet_name(base[: (31 - len(suffix))] + suffix, base)
                        k += 1
                    used_names.add(name)

                    ws = wb.create_sheet(title=name)
                    write_df_to_sheet(ws, df)
                    apply_excel_formatting(ws, df, opts.max_col_width, opts.wrap_text)
                    stats.outputs_created += 1
                    self.log.emit(f"OK: added sheet '{name}'")

                except Exception as e:
                    stats.errors += 1
                    msg = f"ERROR: {csv_path} | {e}"
                    self.log.emit(msg)
                    self.error_item.emit(msg)

            if not self._cancel:
                wb.save(out_path)
                self.log.emit(f"OK: saved {out_path}")
            return

        for i, csv_path in enumerate(self.paths, start=1):
            if self._cancel:
                self.log.emit("Cancelled by user.")
                break

            self.progress.emit(i, total)
            self.status.emit(f"Reading CSV: {csv_path}")

            try:
                df = read_csv_safely(csv_path, opts.delimiter, opts.encoding, opts.dtype_mode)
                wb = Workbook()
                ws = wb.active
                ws.title = "Data"
                write_df_to_sheet(ws, df)
                apply_excel_formatting(ws, df, opts.max_col_width, opts.wrap_text)

                out_path = os.path.join(opts.output_root, f"{base_name_no_ext(csv_path)}.xlsx")
                out_path, should_write = handle_conflict(out_path, opts.conflict_policy)
                if not should_write:
                    stats.skipped += 1
                    self.log.emit(f"SKIP (exists): {out_path}")
                    continue

                wb.save(out_path)
                stats.outputs_created += 1
                self.log.emit(f"OK: {out_path}")

            except Exception as e:
                stats.errors += 1
                msg = f"ERROR: {csv_path} | {e}"
                self.log.emit(msg)
                self.error_item.emit(msg)


class FileDropList(QListWidget):
    def __init__(self, accept_ext: str):
        super().__init__()
        self.accept_ext = accept_ext.lower()
        self.setAcceptDrops(True)
        self.setDragDropMode(QListWidget.DragDropMode.DropOnly)
        self.setSelectionMode(QListWidget.SelectionMode.ExtendedSelection)

    def dragEnterEvent(self, event: QDragEnterEvent):
        if event.mimeData().hasUrls():
            event.acceptProposedAction()
        else:
            super().dragEnterEvent(event)

    def dropEvent(self, event: QDropEvent):
        if not event.mimeData().hasUrls():
            return
        for url in event.mimeData().urls():
            path = url.toLocalFile()
            if path and path.lower().endswith(self.accept_ext):
                self._add_unique(path)
        event.acceptProposedAction()

    def _add_unique(self, path: str):
        for i in range(self.count()):
            if self.item(i).text() == path:
                return
        self.addItem(QListWidgetItem(path))


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Excel ↔ CSV Converter (PyQt6) — Progress + Stats")
        self.resize(1180, 760)

        self.settings = QSettings("PrashantTools", "ExcelCsvConverter")
        self.worker: Optional[ConvertWorker] = None

        self.tabs = QTabWidget()
        self.tabs.addTab(self._build_xlsx_to_csv_tab(), "XLSX → CSV")
        self.tabs.addTab(self._build_csv_to_xlsx_tab(), "CSV → XLSX")

        main = QWidget()
        main_layout = QVBoxLayout(main)
        main_layout.addWidget(self.tabs)

        prog_box = QGroupBox("Progress & Stats")
        prog_layout = QVBoxLayout(prog_box)

        row1 = QHBoxLayout()
        self.progress_bar = QProgressBar()
        self.progress_bar.setRange(0, 100)
        self.progress_label = QLabel("Idle")
        row1.addWidget(self.progress_bar, 2)
        row1.addWidget(self.progress_label, 3)
        prog_layout.addLayout(row1)

        row2 = QHBoxLayout()
        self.stats_label = QLabel("Inputs: 0 | Outputs: 0 | Skipped: 0 | Errors: 0 | Duration: 0.0s")
        self.cancel_btn = QPushButton("Cancel")
        self.cancel_btn.setProperty("btnRole", "danger")
        self.cancel_btn.setEnabled(False)
        self.cancel_btn.clicked.connect(self._cancel_run)
        row2.addWidget(self.stats_label, 1)
        row2.addWidget(self.cancel_btn)
        prog_layout.addLayout(row2)

        main_layout.addWidget(prog_box)

        log_box = QGroupBox("Log")
        log_layout = QVBoxLayout(log_box)
        self.log_text = QPlainTextEdit()
        self.log_text.setReadOnly(True)
        log_layout.addWidget(self.log_text)
        main_layout.addWidget(log_box, 2)

        err_box = QGroupBox("Errors (click to copy)")
        err_layout = QVBoxLayout(err_box)
        self.err_list = QListWidget()
        self.err_list.itemClicked.connect(lambda it: QApplication.clipboard().setText(it.text()))
        err_layout.addWidget(self.err_list)
        main_layout.addWidget(err_box, 1)

        self.setCentralWidget(main)

        self._apply_pastel_theme()  # (2) pastel buttons
        self._load_settings()

    def _apply_pastel_theme(self):
        # Pastel, readable, with hover/pressed states.
        # Roles: primary / secondary / danger
        self.setStyleSheet("""
            QWidget { font-size: 12px; }
            QGroupBox { font-weight: 600; }
            QPushButton {
                border: 1px solid rgba(0,0,0,0.10);
                border-radius: 10px;
                padding: 7px 12px;
                background: #F3F4F6;   /* default pastel gray */
            }
            QPushButton:hover { background: #ECEFF3; }
            QPushButton:pressed { background: #E2E6EC; }
            QPushButton:disabled { opacity: 0.55; }

            QPushButton[btnRole="primary"] {
                background: #DCEBFF;  /* pastel blue */
            }
            QPushButton[btnRole="primary"]:hover { background: #D3E5FF; }
            QPushButton[btnRole="primary"]:pressed { background: #C7DCFF; }

            QPushButton[btnRole="secondary"] {
                background: #DCF7E6;  /* pastel green */
            }
            QPushButton[btnRole="secondary"]:hover { background: #D1F3DF; }
            QPushButton[btnRole="secondary"]:pressed { background: #C3EFD6; }

            QPushButton[btnRole="danger"] {
                background: #FFE0E3;  /* pastel red */
            }
            QPushButton[btnRole="danger"]:hover { background: #FFD6DA; }
            QPushButton[btnRole="danger"]:pressed { background: #FFC9CF; }

            QProgressBar { border: 1px solid rgba(0,0,0,0.15); border-radius: 10px; height: 18px; }
            QProgressBar::chunk { border-radius: 10px; background: #CFE9FF; } /* pastel sky */
        """)

    # ----------------- Settings -----------------
    def _load_settings(self):
        self.xlsx_out_folder.setText(self.settings.value("xlsx_out_folder", ""))
        self.xlsx_mode.setCurrentText(self.settings.value("xlsx_mode", "Single folder"))
        self.xlsx_delim.setCurrentText(self.settings.value("xlsx_delim", "Comma (,)"))
        self.xlsx_encoding.setCurrentText(self.settings.value("xlsx_encoding", "utf-8"))
        self.xlsx_include_index.setChecked(self.settings.value("xlsx_include_index", "false") == "true")
        self.xlsx_first_row_header.setChecked(self.settings.value("xlsx_first_row_header", "true") == "true")
        self.xlsx_hidden_sheets.setChecked(self.settings.value("xlsx_hidden_sheets", "false") == "true")
        self.xlsx_conflict.setCurrentText(self.settings.value("xlsx_conflict", "Rename (add _1, _2 ...)"))

        self.csv_out_folder.setText(self.settings.value("csv_out_folder", ""))
        self.csv_mode.setCurrentText(self.settings.value("csv_mode", "One workbook per CSV"))
        self.merged_name.setText(self.settings.value("merged_name", "merged.xlsx"))
        self.csv_delim.setCurrentText(self.settings.value("csv_delim", "Comma (,)"))
        self.csv_encoding.setCurrentText(self.settings.value("csv_encoding", "utf-8"))
        self.dtype_mode.setCurrentText(self.settings.value("dtype_mode", "Infer types (recommended)"))
        self.wrap_text.setChecked(self.settings.value("wrap_text", "true") == "true")
        self.max_width.setValue(int(self.settings.value("max_width", 60)))
        self.csv_conflict.setCurrentText(self.settings.value("csv_conflict", "Rename (add _1, _2 ...)"))

    def _save_settings(self):
        self.settings.setValue("xlsx_out_folder", self.xlsx_out_folder.text().strip())
        self.settings.setValue("xlsx_mode", self.xlsx_mode.currentText())
        self.settings.setValue("xlsx_delim", self.xlsx_delim.currentText())
        self.settings.setValue("xlsx_encoding", self.xlsx_encoding.currentText())
        self.settings.setValue("xlsx_include_index", "true" if self.xlsx_include_index.isChecked() else "false")
        self.settings.setValue("xlsx_first_row_header", "true" if self.xlsx_first_row_header.isChecked() else "false")
        self.settings.setValue("xlsx_hidden_sheets", "true" if self.xlsx_hidden_sheets.isChecked() else "false")
        self.settings.setValue("xlsx_conflict", self.xlsx_conflict.currentText())

        self.settings.setValue("csv_out_folder", self.csv_out_folder.text().strip())
        self.settings.setValue("csv_mode", self.csv_mode.currentText())
        self.settings.setValue("merged_name", self.merged_name.text().strip())
        self.settings.setValue("csv_delim", self.csv_delim.currentText())
        self.settings.setValue("csv_encoding", self.csv_encoding.currentText())
        self.settings.setValue("dtype_mode", self.dtype_mode.currentText())
        self.settings.setValue("wrap_text", "true" if self.wrap_text.isChecked() else "false")
        self.settings.setValue("max_width", self.max_width.value())
        self.settings.setValue("csv_conflict", self.csv_conflict.currentText())

    # ----------------- UI -----------------
    def _build_xlsx_to_csv_tab(self) -> QWidget:
        root = QWidget()
        layout = QVBoxLayout(root)

        file_box = QGroupBox("Input XLSX files (Drag & drop supported)")
        file_layout = QVBoxLayout(file_box)
        self.xlsx_list = FileDropList(".xlsx")
        file_layout.addWidget(self.xlsx_list)

        btn_row = QHBoxLayout()
        add_btn = QPushButton("Add .xlsx…"); add_btn.setProperty("btnRole", "secondary")
        rem_btn = QPushButton("Remove selected"); rem_btn.setProperty("btnRole", "danger")
        clr_btn = QPushButton("Clear"); clr_btn.setProperty("btnRole", "danger")
        btn_row.addWidget(add_btn); btn_row.addWidget(rem_btn); btn_row.addWidget(clr_btn)
        btn_row.addStretch(1)
        file_layout.addLayout(btn_row)

        add_btn.clicked.connect(self._add_xlsx_files)
        rem_btn.clicked.connect(lambda: self._remove_selected(self.xlsx_list))
        clr_btn.clicked.connect(self.xlsx_list.clear)

        layout.addWidget(file_box)

        opt_box = QGroupBox("Output options")
        opt_layout = QVBoxLayout(opt_box)

        out_row = QHBoxLayout()
        out_row.addWidget(QLabel("Output folder:"))
        self.xlsx_out_folder = QLineEdit()
        browse_out = QPushButton("Browse…"); browse_out.setProperty("btnRole", "primary")
        open_out = QPushButton("Open folder"); open_out.setProperty("btnRole", "secondary")
        out_row.addWidget(self.xlsx_out_folder, 1)
        out_row.addWidget(browse_out)
        out_row.addWidget(open_out)
        opt_layout.addLayout(out_row)

        browse_out.clicked.connect(lambda: self._pick_output_folder(self.xlsx_out_folder))
        open_out.clicked.connect(lambda: self._open_folder(self.xlsx_out_folder.text().strip()))

        row = QHBoxLayout()
        row.addWidget(QLabel("Output mode:"))
        self.xlsx_mode = QComboBox()
        self.xlsx_mode.addItems(["Single folder", "Per workbook folder"])
        row.addWidget(self.xlsx_mode)

        row.addWidget(QLabel("Delimiter:"))
        self.xlsx_delim = QComboBox()
        self.xlsx_delim.addItems(["Comma (,)", "Tab (\\t)", "Semicolon (;)", "Pipe (|)"])
        row.addWidget(self.xlsx_delim)

        row.addWidget(QLabel("Encoding:"))
        self.xlsx_encoding = QComboBox()
        self.xlsx_encoding.addItems(["utf-8", "utf-8-sig", "cp1252"])
        row.addWidget(self.xlsx_encoding)

        row.addWidget(QLabel("Conflict:"))
        self.xlsx_conflict = QComboBox()
        self.xlsx_conflict.addItems(CONFLICT_POLICIES)
        row.addWidget(self.xlsx_conflict)

        row.addStretch(1)
        opt_layout.addLayout(row)

        row2 = QHBoxLayout()
        self.xlsx_include_index = QCheckBox("Include index column in CSV")
        self.xlsx_first_row_header = QCheckBox("First row is header"); self.xlsx_first_row_header.setChecked(True)
        self.xlsx_hidden_sheets = QCheckBox("Include hidden sheets")
        row2.addWidget(self.xlsx_include_index)
        row2.addWidget(self.xlsx_first_row_header)
        row2.addWidget(self.xlsx_hidden_sheets)
        row2.addStretch(1)
        opt_layout.addLayout(row2)

        run_row = QHBoxLayout()
        self.xlsx_run_btn = QPushButton("Convert XLSX → CSV"); self.xlsx_run_btn.setProperty("btnRole", "primary")
        self.xlsx_run_btn.clicked.connect(self._run_xlsx_to_csv)
        run_row.addWidget(self.xlsx_run_btn)
        run_row.addStretch(1)
        opt_layout.addLayout(run_row)

        layout.addWidget(opt_box)
        return root

    def _build_csv_to_xlsx_tab(self) -> QWidget:
        root = QWidget()
        layout = QVBoxLayout(root)

        file_box = QGroupBox("Input CSV files (Drag & drop supported)")
        file_layout = QVBoxLayout(file_box)
        self.csv_list = FileDropList(".csv")
        file_layout.addWidget(self.csv_list)

        btn_row = QHBoxLayout()
        add_btn = QPushButton("Add .csv…"); add_btn.setProperty("btnRole", "secondary")
        rem_btn = QPushButton("Remove selected"); rem_btn.setProperty("btnRole", "danger")
        clr_btn = QPushButton("Clear"); clr_btn.setProperty("btnRole", "danger")
        btn_row.addWidget(add_btn); btn_row.addWidget(rem_btn); btn_row.addWidget(clr_btn)
        btn_row.addStretch(1)
        file_layout.addLayout(btn_row)

        add_btn.clicked.connect(self._add_csv_files)
        rem_btn.clicked.connect(lambda: self._remove_selected(self.csv_list))
        clr_btn.clicked.connect(self.csv_list.clear)

        layout.addWidget(file_box)

        opt_box = QGroupBox("Output options (Well-formatted XLSX)")
        opt_layout = QVBoxLayout(opt_box)

        out_row = QHBoxLayout()
        out_row.addWidget(QLabel("Output folder:"))
        self.csv_out_folder = QLineEdit()
        browse_out = QPushButton("Browse…"); browse_out.setProperty("btnRole", "primary")
        open_out = QPushButton("Open folder"); open_out.setProperty("btnRole", "secondary")
        out_row.addWidget(self.csv_out_folder, 1)
        out_row.addWidget(browse_out)
        out_row.addWidget(open_out)
        opt_layout.addLayout(out_row)

        browse_out.clicked.connect(lambda: self._pick_output_folder(self.csv_out_folder))
        open_out.clicked.connect(lambda: self._open_folder(self.csv_out_folder.text().strip()))

        mode_row = QHBoxLayout()
        mode_row.addWidget(QLabel("Output mode:"))
        self.csv_mode = QComboBox()
        self.csv_mode.addItems(["One workbook per CSV", "Merge into one workbook"])
        mode_row.addWidget(self.csv_mode)

        mode_row.addWidget(QLabel("Merged workbook name:"))
        self.merged_name = QLineEdit("merged.xlsx")
        mode_row.addWidget(self.merged_name, 1)

        mode_row.addWidget(QLabel("Conflict:"))
        self.csv_conflict = QComboBox()
        self.csv_conflict.addItems(CONFLICT_POLICIES)
        mode_row.addWidget(self.csv_conflict)
        opt_layout.addLayout(mode_row)

        fmt_row = QHBoxLayout()
        fmt_row.addWidget(QLabel("Delimiter:"))
        self.csv_delim = QComboBox()
        self.csv_delim.addItems(["Comma (,)", "Tab (\\t)", "Semicolon (;)", "Pipe (|)"])
        fmt_row.addWidget(self.csv_delim)

        fmt_row.addWidget(QLabel("Encoding:"))
        self.csv_encoding = QComboBox()
        self.csv_encoding.addItems(["utf-8", "utf-8-sig", "cp1252"])
        fmt_row.addWidget(self.csv_encoding)

        fmt_row.addWidget(QLabel("Types:"))
        self.dtype_mode = QComboBox()
        self.dtype_mode.addItems(["Infer types (recommended)", "All as text"])
        fmt_row.addWidget(self.dtype_mode)

        self.wrap_text = QCheckBox("Wrap text"); self.wrap_text.setChecked(True)
        fmt_row.addWidget(self.wrap_text)

        fmt_row.addWidget(QLabel("Max col width:"))
        self.max_width = QSpinBox()
        self.max_width.setRange(20, 120)
        self.max_width.setValue(60)
        fmt_row.addWidget(self.max_width)

        fmt_row.addStretch(1)
        opt_layout.addLayout(fmt_row)

        run_row = QHBoxLayout()
        self.csv_run_btn = QPushButton("Convert CSV → XLSX"); self.csv_run_btn.setProperty("btnRole", "primary")
        self.csv_run_btn.clicked.connect(self._run_csv_to_xlsx)
        run_row.addWidget(self.csv_run_btn)
        run_row.addStretch(1)
        opt_layout.addLayout(run_row)

        layout.addWidget(opt_box)
        return root

    # ----------------- Common UI actions -----------------
    def _add_xlsx_files(self):
        paths, _ = QFileDialog.getOpenFileNames(self, "Select XLSX files", filter="Excel (*.xlsx)")
        for p in paths:
            if p and is_xlsx(p):
                self.xlsx_list._add_unique(p)

    def _add_csv_files(self):
        paths, _ = QFileDialog.getOpenFileNames(self, "Select CSV files", filter="CSV (*.csv)")
        for p in paths:
            if p and is_csv(p):
                self.csv_list._add_unique(p)

    def _remove_selected(self, lw: QListWidget):
        for item in lw.selectedItems():
            lw.takeItem(lw.row(item))

    def _pick_output_folder(self, line_edit: QLineEdit):
        folder = QFileDialog.getExistingDirectory(self, "Select output folder")
        if folder:
            line_edit.setText(folder)

    def _open_folder(self, folder: str):
        if not folder:
            return
        ensure_dir(folder)
        if sys.platform.startswith("win"):
            os.startfile(folder)  # noqa
        else:
            try:
                import subprocess
                subprocess.Popen(["xdg-open", folder])
            except Exception:
                pass

    # ----------------- Progress / logs / stats -----------------
    def _append_log(self, text: str):
        self.log_text.appendPlainText(text)

    def _add_error(self, text: str):
        self.err_list.addItem(QListWidgetItem(text))

    def _set_progress(self, cur: int, total: int):
        if total <= 0:
            self.progress_bar.setValue(0)
            return
        self.progress_bar.setValue(int((cur / total) * 100))

    def _set_status(self, text: str):
        self.progress_label.setText(text)

    def _set_stats(self, stats: RunStats):
        self.stats_label.setText(
            f"Inputs: {stats.inputs} | Outputs: {stats.outputs_created} | "
            f"Skipped: {stats.skipped} | Errors: {stats.errors} | "
            f"Duration: {stats.duration_sec():.1f}s"
        )

    def _start_run_ui(self):
        self.cancel_btn.setEnabled(True)
        self.xlsx_run_btn.setEnabled(False)
        self.csv_run_btn.setEnabled(False)
        self.tabs.setEnabled(False)

    def _end_run_ui(self):
        self.cancel_btn.setEnabled(False)
        self.xlsx_run_btn.setEnabled(True)
        self.csv_run_btn.setEnabled(True)
        self.tabs.setEnabled(True)

    def _cancel_run(self):
        if self.worker:
            self.worker.cancel()
            self._append_log("Cancel requested…")

    def _finish_run(self, stats: RunStats):
        self.worker = None
        self._set_progress(100, 100)
        self._set_status("Done")
        self._set_stats(stats)
        self._append_log(f"--- DONE in {stats.duration_sec():.1f}s ---")
        self._end_run_ui()
        self._save_settings()

    # ----------------- Run conversions -----------------
    def _run_xlsx_to_csv(self):
        paths = [self.xlsx_list.item(i).text() for i in range(self.xlsx_list.count())]
        if not paths:
            QMessageBox.warning(self, "No input", "Please add at least one .xlsx file.")
            return

        out_root = self.xlsx_out_folder.text().strip()
        if not out_root:
            QMessageBox.warning(self, "No output folder", "Please choose an output folder.")
            return

        ensure_dir(out_root)

        opts = XlsxToCsvOptions(
            output_root=out_root,
            output_mode=self.xlsx_mode.currentText(),
            delimiter=guess_delimiter(self.xlsx_delim.currentText()),
            encoding=self.xlsx_encoding.currentText(),
            include_index=self.xlsx_include_index.isChecked(),
            first_row_header=self.xlsx_first_row_header.isChecked(),
            include_hidden_sheets=self.xlsx_hidden_sheets.isChecked(),
            conflict_policy=self.xlsx_conflict.currentText(),
        )

        self.err_list.clear()
        self.log_text.clear()
        self.progress_bar.setValue(0)
        self._set_status("Starting…")
        self._append_log("--- XLSX → CSV ---")
        self._append_log(f"Inputs: {len(paths)}")
        self._append_log(f"Output root: {opts.output_root} | Mode: {opts.output_mode}")

        self.worker = ConvertWorker("xlsx_to_csv", paths, x2c=opts, c2x=None)
        self.worker.progress.connect(self._set_progress)
        self.worker.status.connect(self._set_status)
        self.worker.log.connect(self._append_log)
        self.worker.error_item.connect(self._add_error)
        self.worker.done.connect(self._finish_run)

        self._start_run_ui()
        self.worker.start()

    def _run_csv_to_xlsx(self):
        paths = [self.csv_list.item(i).text() for i in range(self.csv_list.count())]
        if not paths:
            QMessageBox.warning(self, "No input", "Please add at least one .csv file.")
            return

        out_root = self.csv_out_folder.text().strip()
        if not out_root:
            QMessageBox.warning(self, "No output folder", "Please choose an output folder.")
            return

        ensure_dir(out_root)

        merged_name = (self.merged_name.text().strip() or "merged.xlsx")
        if not merged_name.lower().endswith(".xlsx"):
            merged_name += ".xlsx"

        opts = CsvToXlsxOptions(
            output_root=out_root,
            output_mode=self.csv_mode.currentText(),
            merged_workbook_name=merged_name,
            delimiter=guess_delimiter(self.csv_delim.currentText()),
            encoding=self.csv_encoding.currentText(),
            dtype_mode=self.dtype_mode.currentText(),
            max_col_width=int(self.max_width.value()),
            wrap_text=self.wrap_text.isChecked(),
            conflict_policy=self.csv_conflict.currentText(),
        )

        self.err_list.clear()
        self.log_text.clear()
        self.progress_bar.setValue(0)
        self._set_status("Starting…")
        self._append_log("--- CSV → XLSX ---")
        self._append_log(f"Inputs: {len(paths)}")
        self._append_log(f"Output root: {opts.output_root} | Mode: {opts.output_mode}")

        self.worker = ConvertWorker("csv_to_xlsx", paths, x2c=None, c2x=opts)
        self.worker.progress.connect(self._set_progress)
        self.worker.status.connect(self._set_status)
        self.worker.log.connect(self._append_log)
        self.worker.error_item.connect(self._add_error)
        self.worker.done.connect(self._finish_run)

        self._start_run_ui()
        self.worker.start()

    def closeEvent(self, event):
        self._save_settings()
        super().closeEvent(event)


def main():
    app = QApplication(sys.argv)
    w = MainWindow()
    w.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
